#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ZIP_CITY_RE = re.compile(r"(?P<zip>\d{3}\s?\d{2})\s*(?P<city>[^\d,][^,]*)$")
TENENET_PREFIX_RE = re.compile(r"^TENENET o\.z\.,\s*", re.IGNORECASE)

PROGRAM_CODE_ALIASES = {
    "AKP": ["AKP_DETI", "AKP_DOSP"],
    "AVL": ["AVL"],
    "KC": ["KC"],
    "KALIA": ["NAS_A_VAZ"],
    "SCPP": ["SCPP"],
    "SPODASK": ["SPODASK"],
    "SVI": ["VCI"],
    "ZZ": ["ZDRAV_ZNEV"],
}

CENTER_ADDRESS_HINTS = {
    "Senec": ("oravská", "lichnerova", "senec"),
    "Banská Bystrica": ("banská bystrica", "bakossova", "kapitulská"),
    "Košice": ("košice", "rastislavova"),
}


def normalize_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def normalize_program_text(value):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def normalize_token(value):
    text = normalize_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()


def split_program_tokens(raw_value):
    tokens = []
    for chunk in normalize_program_text(raw_value).replace(",", "\n").split("\n"):
        token = chunk.strip()
        if token:
            tokens.append(token)
    return tokens


def parse_address(raw_address):
    cleaned = normalize_text(raw_address)
    cleaned = TENENET_PREFIX_RE.sub("", cleaned)
    cleaned = re.sub(r"(\d{3}\s?\d{2})([A-Za-zÁÄČĎÉÍĹĽŇÓÔŔŠŤÚÝŽáäčďéíĺľňóôŕšťúýž])", r"\1 \2", cleaned)

    zip_code = ""
    city = ""
    street = cleaned

    match = ZIP_CITY_RE.search(cleaned)
    if match:
        zip_code = match.group("zip").strip()
        city = match.group("city").strip(" ,")
        street = cleaned[: match.start()].strip(" ,")
    else:
        parts = [part.strip() for part in cleaned.split(",") if part.strip()]
        if len(parts) >= 2:
            city = parts[-1]
            street = ", ".join(parts[:-1])

    name = cleaned[:255]
    return {
        "name": name,
        "street": street[:255] if street else False,
        "zip": zip_code or False,
        "city": city[:255] if city else False,
    }


def load_program_contacts(workbook):
    ws = workbook["Programy "]
    center = None
    contacts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_center, _program, code, person, _leaflet, email, phone = row[:7]
        if row_center:
            center = normalize_text(row_center)
        if not center or not code:
            continue
        contacts[(center, normalize_token(code))] = {
            "responsible_name": normalize_text(person),
            "email": normalize_text(email),
            "phone": normalize_text(phone),
        }
    return contacts


def infer_center(address_text, city):
    haystack = f"{normalize_text(address_text)} {normalize_text(city)}".lower()
    for center, hints in CENTER_ADDRESS_HINTS.items():
        if any(hint in haystack for hint in hints):
            return center
    return None


def resolve_contact(contact_map, center, raw_program_text):
    if not center:
        return {}
    for token in split_program_tokens(raw_program_text):
        contact = contact_map.get((center, normalize_token(token)))
        if contact:
            return {
                "email": contact["email"] or False,
                "phone": contact["phone"] or False,
            }
    return {}


def resolve_program_ids(env, raw_program_text):
    program_model = env["tenenet.program"].with_context(active_test=False)
    all_programs = {program.code: program.id for program in program_model.search([])}
    matched_ids = []
    unmatched = []
    tokens = split_program_tokens(raw_program_text)

    for token in tokens:
        norm_token = normalize_token(token)
        if norm_token.startswith("DETTO AKO VYSSIE"):
            unmatched.append(token)
            continue
        mapped_codes = PROGRAM_CODE_ALIASES.get(norm_token)
        if not mapped_codes:
            unmatched.append(token)
            continue
        for code in mapped_codes:
            program_id = all_programs.get(code)
            if program_id and program_id not in matched_ids:
                matched_ids.append(program_id)
    return matched_ids, unmatched


def get_landlord(env, raw_name):
    landlord_name = normalize_text(raw_name).rstrip(",")
    if not landlord_name or landlord_name == "0" or landlord_name == "ZVÝŠENÉ":
        return env["res.partner"]
    Partner = env["res.partner"].with_context(active_test=False)
    landlord = Partner.search(
        [("name", "=", landlord_name), ("is_tenenet_landlord", "=", True)],
        limit=1,
    )
    if landlord:
        return landlord
    landlord = Partner.search([("name", "=", landlord_name)], limit=1)
    if landlord:
        landlord.write({"is_tenenet_landlord": True})
        return landlord
    return Partner.create(
        {
            "name": landlord_name,
            "is_tenenet_landlord": True,
            "company_type": "company",
        }
    )


def import_sites(env, workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Hárok1"]
    contact_map = load_program_contacts(wb)

    Country = env["res.country"]
    slovakia = Country.search([("code", "=", "SK")], limit=1)
    Site = env["tenenet.project.site"].with_context(active_test=False)

    previous_program_text = ""
    imported = 0
    updated = 0
    skipped = []
    unmatched_summary = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        kraj, raw_programs, _usage, raw_address, raw_landlord, *_rest = row
        address_text = normalize_text(raw_address)
        if not address_text:
            skipped.append((row_idx, "missing address"))
            continue

        program_text = normalize_program_text(raw_programs)
        if normalize_token(program_text).startswith("DETTO AKO VYSSIE"):
            program_text = previous_program_text
        elif program_text:
            previous_program_text = program_text

        program_ids, unmatched = resolve_program_ids(env, program_text)
        if unmatched:
            unmatched_summary.append((row_idx, address_text, unmatched))

        landlord = get_landlord(env, raw_landlord)
        address_vals = parse_address(address_text)
        center = infer_center(address_text, address_vals["city"])
        contact_vals = resolve_contact(contact_map, center, program_text)
        vals = {
            **address_vals,
            **contact_vals,
            "site_type": "prevadzka",
            "kraj": env["tenenet.project.site"]._normalize_region_label(normalize_text(kraj)) or False,
            "country_id": slovakia.id if slovakia else False,
            "landlord_partner_id": landlord.id or False,
            "program_ids": [(6, 0, program_ids)],
            "legacy_program_text": program_text or False,
        }

        domain = [("site_type", "=", "prevadzka"), ("name", "=", vals["name"])]
        existing = Site.search(domain, limit=1)
        if not existing and vals["street"] and vals["city"]:
            existing = Site.search(
                [
                    ("site_type", "=", "prevadzka"),
                    ("street", "=", vals["street"]),
                    ("city", "=", vals["city"]),
                ],
                limit=1,
            )

        if existing:
            existing.write(vals)
            updated += 1
        else:
            Site.create(vals)
            imported += 1

    return {
        "created": imported,
        "updated": updated,
        "skipped": skipped,
        "unmatched": unmatched_summary,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    result = import_sites(env, workbook_path)
    env.cr.commit()
    print(f"CREATED={result['created']}")
    print(f"UPDATED={result['updated']}")
    if result["skipped"]:
        print("SKIPPED_ROWS:")
        for row_idx, reason in result["skipped"]:
            print(f"  row {row_idx}: {reason}")
    if result["unmatched"]:
        print("UNMATCHED_PROGRAM_TOKENS:")
        for row_idx, address, tokens in result["unmatched"]:
            print(f"  row {row_idx} [{address}]: {', '.join(tokens)}")


main()
