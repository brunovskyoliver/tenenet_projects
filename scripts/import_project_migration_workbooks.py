#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from calendar import monthrange
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PROJECT_SUMMARY_SHEET = "Projects Summary MG"
ASSIGNMENT_AGGREGATE_SHEETS = {"summary_Tenenet"}
CCP_MULTIPLIER = 1.362
YEAR = 2026
DEFAULT_CASHFLOW_SHEET = "CF 2026 (rolling)"
CASHFLOW_SHEET_FALLBACK = "CF 2026"
MONTH_COLUMNS = (3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15)
MONTH_BY_COLUMN = {
    3: 1,
    4: 2,
    5: 3,
    6: 4,
    7: 5,
    8: 6,
    10: 7,
    11: 8,
    12: 9,
    13: 10,
    14: 11,
    15: 12,
}
CF_MONTH_COLUMNS = {
    22: 1,
    23: 2,
    24: 3,
    25: 4,
    26: 5,
    27: 6,
    28: 7,
    29: 8,
    30: 9,
    31: 10,
    32: 11,
    33: 12,
}
CF_PLAN_INCOME_ROWS = range(3, 48)
CF_PLAN_EXPENSE_ROWS = range(50, 79)
CF_PLAN_SUBTOTAL_LABELS = {
    "cash-in, bank statement",
    "cash-out bank statement",
}
APZ_PROJECT_NAME = "APZ_2N"
APZ_SPECIAL_MONTH_COLUMNS = {
    9: 1,
    10: 2,
    11: 3,
    12: 4,
    13: 5,
    14: 6,
    15: 7,
    16: 8,
    17: 9,
    18: 10,
    19: 11,
    20: 12,
}

PROGRAM_CODE_ALIASES = {
    "admin": "ADMIN_TENENET",
    "admin tenenet": "ADMIN_TENENET",
    "kalia nasilie": "NAS_A_VAZ",
    "kc": "KC",
    "scpp": "SCPP",
    "spodask": "SPODASK",
    "spodask cdr": "SPODASK",
    "nasilie a vazenstvo": "NAS_A_VAZ",
    "svi": "VCI",
    "vci": "VCI",
    "ssp": "SSP",
    "šsp": "SSP",
    "šsp": "SSP",
    "zz": "ZDRAV_ZNEV",
}

PROJECT_TYPE_ALIASES = {
    "1 sr - samosprava": "narodny",
    "2 esf - dopytovy": "narodny",
    "3 esf - narodny": "narodny",
    "03 esf narodny": "narodny",
    "10 plan obnovy a odolnosti": "narodny",
    "4 erasmus": "medzinarodny",
    "8 eu": "medzinarodny",
    "08 eu": "medzinarodny",
    "11 vysegrad fund": "medzinarodny",
    "narodny": "narodny",
    "medzinarodny": "medzinarodny",
    "sluzby": "sluzby",
}

NON_PROJECT_CASHFLOW_PATTERNS = (
    "balance",
    "opening balance",
    "cash-out",
    "cash out",
    "prevadzkove n",
    "investicne n",
    "financny n",
    "vklad z pokladne",
)
EMPLOYEE_TITLE_RE = re.compile(
    r"(?<!\w)(?:bc|doc|ing|judr|mga|mgr|mudr|paeddr|phd|phdr|prof|rndr)\.?(?!\w)",
    re.IGNORECASE,
)


@dataclass
class MigrationWarning:
    source: str
    row: int | str
    level: str
    message: str
    raw_value: Any = ""


@dataclass
class UnmatchedRow:
    source: str
    row: int | str
    kind: str
    key: str
    reason: str
    context: str = ""


@dataclass
class EmployeeMatch:
    name: str
    employee_id: int | bool = False
    user_id: int | bool = False
    user_login: str = ""
    user_email: str = ""
    work_email: str = ""
    active: bool | str = ""
    match_status: str = "source_workbook"
    source_ref: str = ""


@dataclass
class EmployeeMonthlySummaryPreview:
    sheet: str
    employee_name: str
    employee_id: int | bool
    employee_match_status: str
    employee_source_ref: str
    capacity_hours_incl: dict[int, float] = field(default_factory=dict)
    capacity_hours: dict[int, float] = field(default_factory=dict)
    total_gross_salary: dict[int, float] = field(default_factory=dict)
    total_labor_cost: dict[int, float] = field(default_factory=dict)
    worked_hours: dict[int, float] = field(default_factory=dict)
    holidays_hours: dict[int, float] = field(default_factory=dict)
    vacation_hours: dict[int, float] = field(default_factory=dict)
    doctor_hours: dict[int, float] = field(default_factory=dict)
    internal_gross_salary: dict[int, float] = field(default_factory=dict)
    internal_labor_cost: dict[int, float] = field(default_factory=dict)


@dataclass
class ProjectProgramBinding:
    project_name: str
    program_raw: str = ""
    program_code: str = ""
    source_ref: str = ""


@dataclass
class ProjectPreview:
    source_row: int
    project_name: str
    description: str
    contract_number: str
    recipient_text: str
    recipient_id: int | bool
    donor_text: str
    donor_id: int | bool
    project_type_raw: str
    project_type: str
    program_raw: str
    program_code: str
    program_id: int | bool
    budget_amount: float
    received_2026: float
    date_start: date | None
    date_end: date | None
    existing_project_id: int | bool
    match_status: str


@dataclass
class AssignmentPreview:
    sheet: str
    project_name: str
    project_id: int | bool
    employee_name: str
    employee_id: int | bool
    employee_user_id: int | bool
    employee_user_login: str
    employee_work_email: str
    employee_match_status: str
    employee_source_ref: str
    total_hours: float
    total_ccp: float
    wage_ccp: float
    wage_hm: float
    contribution_multiplier: float
    allocation_ratio: float
    date_start: date
    date_end: date
    monthly_ratios: dict[int, float] = field(default_factory=dict)


@dataclass
class TimesheetPreview:
    sheet: str
    project_name: str
    employee_name: str
    period: date
    hours_pp: float
    expected_hours: float
    allocation_ratio: float
    ccp_amount: float


@dataclass
class CashflowPreview:
    source_row: int
    source_sheet: str
    program_label: str
    project_label: str
    project_id: int | bool
    receipt_date: date
    receipt_amount: float
    receipt_note: str
    month_amounts: dict[int, float] = field(default_factory=dict)


@dataclass
class CashflowPlanPreview:
    source_row: int
    source_sheet: str
    row_key: str
    row_label: str
    row_type: str
    section_label: str
    program_label: str
    sequence: int
    actual_mapping_key: str
    month_amounts: dict[int, float] = field(default_factory=dict)


@dataclass
class ImportAction:
    source: str
    action: str
    model: str
    record_id: int | bool
    key: str
    status: str
    message: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def fold_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", normalize_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def project_key(value: Any) -> str:
    text = fold_text(value)
    text = re.sub(r"^\d+\s+", "", text)
    text = text.replace("_", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def cashflow_row_slug(value: Any) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", project_key(value))
    return text.strip("-") or "row"


def parse_amount(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = normalize_text(value)
    match = re.search(r"-?\d+(?:[ \u00a0.]?\d{3})*(?:,\d+)?|-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    number = match.group(0).replace(" ", "").replace("\xa0", "")
    if "," in number:
        number = number.replace(".", "").replace(",", ".")
    try:
        return float(number)
    except ValueError:
        return 0.0


def parse_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_program_code(raw_value: Any) -> str:
    key = fold_text(raw_value)
    key = re.sub(r"[^a-z0-9š]+", " ", key)
    key = " ".join(key.split())
    return PROGRAM_CODE_ALIASES.get(key, "")


def normalize_project_type(raw_value: Any) -> str:
    key = re.sub(r"[^a-z0-9]+", " ", fold_text(raw_value))
    key = " ".join(key.split())
    return PROJECT_TYPE_ALIASES.get(key, "narodny")


def row_has_2026_scope(row: tuple[Any, ...]) -> bool:
    year_text = normalize_text(row[0])
    if str(YEAR) in year_text:
        return True

    start = parse_date(row[25])
    end = parse_date(row[26])
    year_start = date(YEAR, 1, 1)
    year_end = date(YEAR, 12, 31)
    if start and end and start <= year_end and end >= year_start:
        return True
    if start and not end and start.year <= YEAR:
        return True
    if end and not start and end.year >= YEAR:
        return True

    return bool(parse_amount(row[33]))


def is_non_project_cashflow(label: str, amount: float) -> bool:
    key = fold_text(label)
    return amount <= 0.0 or any(pattern in key for pattern in NON_PROJECT_CASHFLOW_PATTERNS)


def is_cashflow_plan_row(row_idx: int, label: str) -> bool:
    key = fold_text(label)
    if key in CF_PLAN_SUBTOTAL_LABELS:
        return True
    return row_idx in CF_PLAN_INCOME_ROWS or row_idx in CF_PLAN_EXPENSE_ROWS


def cashflow_plan_row_type(row_idx: int, label: str) -> str:
    key = fold_text(label)
    if row_idx in CF_PLAN_EXPENSE_ROWS and key.startswith("mzdy"):
        return "salary"
    if row_idx in CF_PLAN_EXPENSE_ROWS:
        return "expense"
    return "income"


def cashflow_plan_row_key(row_type: str, label: str) -> str:
    return f"workbook:{row_type}:{cashflow_row_slug(label)}"


def employee_aliases(value: str) -> set[str]:
    normalized = normalize_text(value)
    aliases = set()
    variants = {normalized} if normalized else set()
    if "_" in normalized:
        variants.add(normalized.split("_", 1)[1])
    for variant in list(variants):
        without_titles = normalize_text(EMPLOYEE_TITLE_RE.sub(" ", variant))
        if without_titles and without_titles != variant:
            variants.add(without_titles)
    for variant in list(variants):
        without_suffix = re.sub(r"\s+\d+h?$", "", variant, flags=re.IGNORECASE)
        without_suffix = re.sub(r"\s+\d+$", "", without_suffix)
        if without_suffix and without_suffix != variant:
            variants.add(without_suffix)
    for variant in variants:
        key = fold_text(variant)
        if key:
            aliases.add(key)
        parts = variant.split()
        if len(parts) == 2:
            aliases.add(f"{fold_text(parts[1])} {fold_text(parts[0])}")
    return {alias for alias in aliases if alias}


def merge_employee_match(existing: EmployeeMatch, incoming: EmployeeMatch) -> EmployeeMatch:
    if existing.match_status.startswith("source") and incoming.match_status.startswith("odoo"):
        return incoming
    if not existing.employee_id and incoming.employee_id:
        return EmployeeMatch(
            name=incoming.name,
            employee_id=incoming.employee_id,
            user_id=incoming.user_id or existing.user_id,
            user_login=incoming.user_login or existing.user_login,
            user_email=incoming.user_email or existing.user_email,
            work_email=incoming.work_email or existing.work_email,
            active=incoming.active if incoming.active != "" else existing.active,
            match_status=incoming.match_status,
            source_ref="; ".join(part for part in [incoming.source_ref, existing.source_ref] if part),
        )
    if existing.employee_id and not existing.user_id and incoming.user_id:
        return EmployeeMatch(
            name=existing.name,
            employee_id=existing.employee_id,
            user_id=incoming.user_id,
            user_login=incoming.user_login,
            user_email=incoming.user_email,
            work_email=existing.work_email or incoming.work_email,
            active=existing.active,
            match_status=existing.match_status,
            source_ref="; ".join(part for part in [existing.source_ref, incoming.source_ref] if part),
        )
    return existing


def add_employee_index_entry(index: dict[str, EmployeeMatch], match: EmployeeMatch) -> None:
    for alias in employee_aliases(match.name):
        existing = index.get(alias)
        index[alias] = merge_employee_match(existing, match) if existing else match


def record_value(record, field_name: str, default: Any = "") -> Any:
    if not record or field_name not in record._fields:
        return default
    value = record[field_name]
    return default if value in (None, "") else value


def add_source_workbook_employees(index: dict[str, EmployeeMatch], workbook) -> None:
    if "Zoznam zamestnancov" in workbook.sheetnames:
        sheet = workbook["Zoznam zamestnancov"]
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=11, values_only=True), start=2):
            name = normalize_text(row[2] if len(row) >= 3 else "")
            if not name or name == "Zamestnanec":
                continue
            source_no = normalize_text(row[0] if row else "")
            add_employee_index_entry(index, EmployeeMatch(
                name=name,
                match_status="source_workbook",
                source_ref=f"Zoznam zamestnancov:{row_idx}" + (f" #{source_no}" if source_no else ""),
            ))

    ignored_sheets = {
        "Meno Priezvisko - template",
        "Zoznam projektov",
        "Zoznam zamestnancov",
        "Sheet6",
        "uprava macra",
    }
    for sheet_name in workbook.sheetnames:
        if sheet_name in ignored_sheets or sheet_name.startswith("summary_"):
            continue
        add_employee_index_entry(index, EmployeeMatch(
            name=sheet_name,
            match_status="source_workbook_sheet",
            source_ref=f"sheet:{sheet_name}",
        ))


def build_employee_index(env, workbook=None) -> dict[str, EmployeeMatch]:
    result = {}
    if env:
        employees = env["hr.employee"].with_context(active_test=False).search([])
        for employee in employees:
            user = record_value(employee, "user_id", False)
            candidates = [
                employee.name or "",
                getattr(employee, "legal_name", "") or "",
                " ".join(part for part in [getattr(employee, "first_name", "") or "", getattr(employee, "last_name", "") or ""] if part),
                " ".join(part for part in [getattr(employee, "last_name", "") or "", getattr(employee, "first_name", "") or ""] if part),
                record_value(employee, "work_email", ""),
                user.login if user else "",
                user.email if user else "",
            ]
            for candidate in candidates:
                if candidate:
                    add_employee_index_entry(result, EmployeeMatch(
                        name=candidate,
                        employee_id=employee.id,
                        user_id=user.id if user else False,
                        user_login=user.login if user else "",
                        user_email=user.email if user else "",
                        work_email=record_value(employee, "work_email", ""),
                        active=record_value(employee, "active", ""),
                        match_status="odoo",
                        source_ref=f"hr.employee:{employee.id}",
                    ))
        for user in env["res.users"].with_context(active_test=False).search([]):
            candidates = [user.name or "", user.login or "", user.email or ""]
            for candidate in candidates:
                if candidate:
                    add_employee_index_entry(result, EmployeeMatch(
                        name=candidate,
                        user_id=user.id,
                        user_login=user.login or "",
                        user_email=user.email or "",
                        active=record_value(user, "active", ""),
                        match_status="odoo_user",
                        source_ref=f"res.users:{user.id}",
                    ))
    if workbook:
        add_source_workbook_employees(result, workbook)
    return result


def resolve_employee(employee_index: dict[str, EmployeeMatch], name: str):
    matches = [employee_index[alias] for alias in employee_aliases(name) if alias in employee_index]
    for match in matches:
        if match.employee_id and match.match_status == "odoo":
            return match
    for match in matches:
        if match.employee_id:
            return match
    if matches:
        return matches[0]
    return None


def employee_short_keys(value: str) -> set[str]:
    normalized = normalize_text(value)
    parts = normalized.split()
    if not parts:
        return set()
    keys = {fold_text(parts[0])}
    if len(parts) >= 2 and parts[1]:
        keys.add(f"{fold_text(parts[0])} {fold_text(parts[1][0])}")
    return {key for key in keys if key}


def build_unique_employee_short_index(employee_index: dict[str, EmployeeMatch]) -> dict[str, EmployeeMatch]:
    candidates = defaultdict(set)
    matches_by_name = {}
    for match in employee_index.values():
        canonical = normalize_text(match.name)
        if not canonical:
            continue
        matches_by_name[canonical] = match
    for name in matches_by_name:
        for key in employee_short_keys(name):
            candidates[key].add(name)
    return {
        key: matches_by_name[next(iter(names))]
        for key, names in candidates.items()
        if len(names) == 1
    }


def resolve_employee_short(employee_index: dict[str, EmployeeMatch], name: str):
    exact = resolve_employee(employee_index, name)
    if exact:
        return exact
    short_index = build_unique_employee_short_index(employee_index)
    for key in employee_short_keys(name):
        match = short_index.get(key)
        if match:
            return match
    return None


def build_summary_employee_short_index(employee_index: dict[str, EmployeeMatch], workbook, sheet_name: str) -> dict[str, EmployeeMatch]:
    if sheet_name not in workbook.sheetnames:
        return {}
    sheet = workbook[sheet_name]
    candidates = defaultdict(set)
    matches_by_name = {}
    for row_idx in range(2, sheet.max_row):
        row_type = fold_text(sheet.cell(row=row_idx, column=2).value)
        next_type = fold_text(sheet.cell(row=row_idx + 1, column=2).value)
        if "celkova cena prace" not in row_type or "odpracovane hodiny" not in next_type:
            continue
        employee_name = normalize_text(sheet.cell(row=row_idx, column=1).value)
        if not employee_name or fold_text(employee_name) in {"total", "spolu"}:
            continue
        source_match = resolve_employee(employee_index, employee_name)
        matches_by_name[employee_name] = EmployeeMatch(
            name=employee_name,
            employee_id=source_match.employee_id if source_match else False,
            match_status=source_match.match_status if source_match else "source_workbook_summary",
            source_ref=source_match.source_ref if source_match else f"{sheet_name}:{row_idx}",
        )
    for employee_name in matches_by_name:
        for key in employee_short_keys(employee_name):
            candidates[key].add(employee_name)
    return {
        key: matches_by_name[next(iter(names))]
        for key, names in candidates.items()
        if len(names) == 1
    }


def resolve_program(env, program_code: str):
    if not env or not program_code:
        return None
    return env["tenenet.program"].with_context(active_test=False).search([("code", "=", program_code)], limit=1)


def resolve_partner(env, name: str):
    if not env or not name:
        return None
    return env["res.partner"].with_context(active_test=False).search([("name", "=", name)], limit=1)


def resolve_donor(env, name: str):
    if not env or not name:
        return None
    return env["tenenet.donor"].with_context(active_test=False).search([("name", "=", name)], limit=1)


def resolve_project(env, name: str):
    if not env or not name:
        return None
    return env["tenenet.project"].with_context(active_test=False).search([("name", "=", name)], limit=1)


def parse_project_summary(env, workbook_path: Path, warnings: list[MigrationWarning]) -> list[ProjectPreview]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[PROJECT_SUMMARY_SHEET]
    projects = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, max_col=44, values_only=True), start=3):
        if not row_has_2026_scope(row):
            continue
        project_name = normalize_text(row[2])
        if not project_name:
            continue

        program_raw = normalize_text(row[9])
        program_code = normalize_program_code(program_raw)
        program = resolve_program(env, program_code)
        if program_raw and not program_code:
            warnings.append(MigrationWarning("Projects", row_idx, "warning", "Unknown project program", program_raw))
        elif program_code and env and not program:
            warnings.append(MigrationWarning("Projects", row_idx, "warning", "Program code not found in Odoo", program_code))

        project_type_raw = normalize_text(row[7])
        project_type = normalize_project_type(project_type_raw)
        if project_type_raw and fold_text(project_type_raw) not in PROJECT_TYPE_ALIASES:
            # Alias map is intentionally small; only warn when value did not map through known forms.
            key = " ".join(re.sub(r"[^a-z0-9]+", " ", fold_text(project_type_raw)).split())
            if key not in PROJECT_TYPE_ALIASES:
                warnings.append(MigrationWarning("Projects", row_idx, "info", "Project type defaulted to narodny", project_type_raw))

        recipient_text = normalize_text(row[6])
        donor_text = normalize_text(row[8])
        recipient = resolve_partner(env, recipient_text)
        donor = resolve_donor(env, donor_text)
        existing_project = resolve_project(env, project_name)
        projects.append(ProjectPreview(
            source_row=row_idx,
            project_name=project_name,
            description=normalize_text(row[3]),
            contract_number=normalize_text(row[5]),
            recipient_text=recipient_text,
            recipient_id=recipient.id if recipient else False,
            donor_text=donor_text,
            donor_id=donor.id if donor else False,
            project_type_raw=project_type_raw,
            project_type=project_type,
            program_raw=program_raw,
            program_code=program_code,
            program_id=program.id if program else False,
            budget_amount=parse_amount(row[19]),
            received_2026=parse_amount(row[33]),
            date_start=parse_date(row[25]),
            date_end=parse_date(row[26]),
            existing_project_id=existing_project.id if existing_project else False,
            match_status="update" if existing_project else "create_preview",
        ))

    return projects


def build_project_index(projects: list[ProjectPreview]) -> dict[str, ProjectPreview]:
    index = {}
    for project in projects:
        for value in (project.project_name, project.description, project.contract_number):
            key = project_key(value)
            if key:
                index.setdefault(key, project)
    return index


def compact_project_match_key(value: str) -> str:
    return re.sub(r"\d+$", "", project_key(value).replace(" ", ""))


def resolve_project_preview(projects_by_key: dict[str, ProjectPreview], label: str) -> ProjectPreview | None:
    key = project_key(label)
    if not key:
        return None
    exact = projects_by_key.get(key)
    if exact:
        return exact

    compact_label = compact_project_match_key(label)
    if len(compact_label) < 3:
        return None

    candidates = {}
    for candidate_key, project in projects_by_key.items():
        compact_candidate = compact_project_match_key(candidate_key)
        if not compact_candidate or len(compact_candidate) < 3:
            continue
        if compact_candidate.startswith(compact_label) or compact_label.startswith(compact_candidate):
            candidates[project.project_name] = project
    return next(iter(candidates.values())) if len(candidates) == 1 else None


def sheet_project_label(sheet_name: str) -> str:
    raw = sheet_name.removeprefix("summary_")
    raw = raw.replace("_", " ")
    return normalize_text(raw)


def load_assignment_project_bindings(env, workbook_path: Path, warnings: list[MigrationWarning]) -> dict[str, ProjectProgramBinding]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    bindings = {}
    if "Zoznam projektov" not in workbook.sheetnames:
        return bindings
    sheet = workbook["Zoznam projektov"]
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=3, values_only=True), start=2):
        project_name = normalize_text(row[1] if len(row) > 1 else "")
        if not project_name:
            continue
        program_raw = normalize_text(row[2] if len(row) > 2 else "")
        program_code = normalize_program_code(program_raw)
        if program_raw and not program_code:
            warnings.append(MigrationWarning("Project Program Bindings", row_idx, "warning", "Unknown assignment project program", program_raw))
        elif program_code and env and not resolve_program(env, program_code):
            warnings.append(MigrationWarning("Project Program Bindings", row_idx, "warning", "Program code not found in Odoo", program_code))
        bindings.setdefault(project_key(project_name), ProjectProgramBinding(
            project_name=project_name,
            program_raw=program_raw,
            program_code=program_code,
            source_ref=f"Zoznam projektov:{row_idx}",
        ))
    return bindings


def closest_assignment_project_binding(project_label: str, bindings: dict[str, ProjectProgramBinding]) -> ProjectProgramBinding:
    label_key = project_key(project_label)
    if not label_key:
        return ProjectProgramBinding(project_name=project_label)
    if label_key in bindings:
        return bindings[label_key]

    compact_label = label_key.replace(" ", "")
    for alias_key, binding in bindings.items():
        compact_alias = alias_key.replace(" ", "")
        if compact_alias and (compact_label == compact_alias or compact_label in compact_alias or compact_alias in compact_label):
            return binding
    return ProjectProgramBinding(project_name=project_label)


def append_assignment_project_previews(
    env,
    workbook_path: Path,
    project_program_bindings_path: Path,
    projects: list[ProjectPreview],
    warnings: list[MigrationWarning],
) -> list[ProjectPreview]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    projects_by_key = build_project_index(projects)
    bindings = load_assignment_project_bindings(env, project_program_bindings_path, warnings)

    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("summary_") or sheet_name in ASSIGNMENT_AGGREGATE_SHEETS:
            continue
        project_label = sheet_project_label(sheet_name)
        if projects_by_key.get(project_key(project_label)):
            continue

        binding = closest_assignment_project_binding(project_label, bindings)
        project_name = binding.project_name
        program = resolve_program(env, binding.program_code)
        existing_project = resolve_project(env, project_name)
        project_type = "narodny" if binding.program_code else "medzinarodny"
        project = ProjectPreview(
            source_row=sheet_name,
            project_name=project_name,
            description="",
            contract_number=project_label,
            recipient_text="",
            recipient_id=False,
            donor_text="",
            donor_id=False,
            project_type_raw="assignment_program_binding" if binding.program_code else "no program binding",
            project_type=project_type,
            program_raw=binding.program_raw,
            program_code=binding.program_code,
            program_id=program.id if program else False,
            budget_amount=0.0,
            received_2026=0.0,
            date_start=date(YEAR, 1, 1),
            date_end=date(YEAR, 12, 31),
            existing_project_id=existing_project.id if existing_project else False,
            match_status="update_from_assignment_workbook" if existing_project else "create_from_assignment_workbook",
        )
        projects.append(project)
        for value in (project.project_name, project_label, sheet_name.removeprefix("summary_")):
            key = project_key(value)
            if key:
                projects_by_key.setdefault(key, project)
        warnings.append(MigrationWarning(
            "Assignments",
            sheet_name,
            "info",
            "Project missing from project summary; dry-run project candidate created from assignment workbook",
            project_name,
        ))

    return projects


def load_expected_hours(workbook) -> dict[int, float]:
    expected = {month: 0.0 for month in range(1, 13)}
    if "Meno Priezvisko - template" not in workbook.sheetnames:
        return expected
    sheet = workbook["Meno Priezvisko - template"]
    for col_idx, month in MONTH_BY_COLUMN.items():
        expected[month] = parse_amount(sheet.cell(row=4, column=col_idx).value)
    return expected


def read_employee_sheet_multiplier(sheet) -> float | None:
    for row_idx in range(1, min(sheet.max_row, 40) + 1):
        label = fold_text(sheet.cell(row=row_idx, column=2).value)
        if "updatnut podla skutocnosti" not in label:
            continue
        values = [
            parse_amount(sheet.cell(row=row_idx, column=col_idx).value)
            for col_idx in MONTH_BY_COLUMN
        ]
        non_zero = [value for value in values if value > 0.0]
        if non_zero:
            return round(non_zero[0], 4)
    return None


def load_employee_contribution_multipliers(
    workbook,
    employee_index: dict[str, EmployeeMatch],
) -> dict[str, float]:
    ignored_sheets = {
        "Meno Priezvisko - template",
        "Zoznam projektov",
        "Zoznam zamestnancov",
        "Sheet6",
        "uprava macra",
    }
    multipliers: dict[str, float] = {}
    for sheet_name in workbook.sheetnames:
        if sheet_name in ignored_sheets or sheet_name.startswith("summary_"):
            continue
        multiplier = read_employee_sheet_multiplier(workbook[sheet_name])
        if not multiplier:
            continue
        match = resolve_employee(employee_index, sheet_name)
        if match and match.employee_id:
            multipliers[f"employee:{int(match.employee_id)}"] = multiplier
        for alias in employee_aliases(sheet_name):
            multipliers[f"alias:{alias}"] = multiplier
    return multipliers


def read_month_values(sheet, row_idx: int) -> dict[int, float]:
    return {
        month: parse_amount(sheet.cell(row=row_idx, column=col_idx).value)
        for col_idx, month in MONTH_BY_COLUMN.items()
    }


def find_sheet_row(
    sheet,
    *,
    col1_contains: str | None = None,
    col2_contains: str | None = None,
    col1_equals: str | None = None,
    col2_equals: str | None = None,
    max_row: int | None = None,
) -> int | None:
    limit = min(sheet.max_row, max_row or sheet.max_row)
    for row_idx in range(1, limit + 1):
        col1 = fold_text(sheet.cell(row=row_idx, column=1).value)
        col2 = fold_text(sheet.cell(row=row_idx, column=2).value)
        if col1_equals and col1 != col1_equals:
            continue
        if col2_equals and col2 != col2_equals:
            continue
        if col1_contains and col1_contains not in col1:
            continue
        if col2_contains and col2_contains not in col2:
            continue
        return row_idx
    return None


def parse_employee_monthly_summaries(
    env,
    workbook_path: Path,
    unmatched: list[UnmatchedRow],
) -> list[EmployeeMonthlySummaryPreview]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    employee_index = build_employee_index(env, workbook)
    ignored_sheets = {
        "Meno Priezvisko - template",
        "Zoznam projektov",
        "Zoznam zamestnancov",
        "Sheet6",
        "uprava macra",
    }
    summaries: list[EmployeeMonthlySummaryPreview] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in ignored_sheets or sheet_name.startswith("summary_"):
            continue
        sheet = workbook[sheet_name]
        employee_match = resolve_employee(employee_index, sheet_name)
        if not employee_match:
            unmatched.append(UnmatchedRow(
                "Employee monthly summary",
                sheet_name,
                "employee",
                sheet_name,
                "Employee sheet not matched",
            ))
            continue

        capacity_incl_row = find_sheet_row(sheet, col1_equals="hours/month (incl holidays)", max_row=20)
        capacity_row = find_sheet_row(sheet, col1_equals="hours/month", max_row=20)
        total_gross_row = find_sheet_row(sheet, col2_contains="hruba mzda", max_row=20)
        total_labor_row = find_sheet_row(sheet, col2_contains="celkova cena prace", max_row=20)
        worked_row = find_sheet_row(sheet, col2_contains="odpracovane hodiny", max_row=20)
        holidays_row = find_sheet_row(sheet, col2_contains="platene sviatky", max_row=20)
        vacation_row = find_sheet_row(sheet, col2_contains="dovolenka hodiny", max_row=20)
        doctor_row = find_sheet_row(sheet, col2_contains="lekar", max_row=20)
        internal_gross_row = find_sheet_row(sheet, col1_contains="tenenet", col2_contains="hruba mzda")
        internal_labor_row = find_sheet_row(sheet, col1_contains="tenenet", col2_contains="celkova cena prace")

        summary = EmployeeMonthlySummaryPreview(
            sheet=sheet_name,
            employee_name=sheet_name,
            employee_id=employee_match.employee_id if employee_match else False,
            employee_match_status=employee_match.match_status if employee_match else "missing",
            employee_source_ref=employee_match.source_ref if employee_match else "",
            capacity_hours_incl=read_month_values(sheet, capacity_incl_row) if capacity_incl_row else {},
            capacity_hours=read_month_values(sheet, capacity_row) if capacity_row else {},
            total_gross_salary=read_month_values(sheet, total_gross_row) if total_gross_row else {},
            total_labor_cost=read_month_values(sheet, total_labor_row) if total_labor_row else {},
            worked_hours=read_month_values(sheet, worked_row) if worked_row else {},
            holidays_hours=read_month_values(sheet, holidays_row) if holidays_row else {},
            vacation_hours=read_month_values(sheet, vacation_row) if vacation_row else {},
            doctor_hours=read_month_values(sheet, doctor_row) if doctor_row else {},
            internal_gross_salary=read_month_values(sheet, internal_gross_row) if internal_gross_row else {},
            internal_labor_cost=read_month_values(sheet, internal_labor_row) if internal_labor_row else {},
        )
        if any(
            any(abs(value or 0.0) > 0.00001 for value in values.values())
            for values in (
                summary.capacity_hours_incl,
                summary.capacity_hours,
                summary.total_gross_salary,
                summary.total_labor_cost,
                summary.worked_hours,
                summary.holidays_hours,
                summary.vacation_hours,
                summary.doctor_hours,
                summary.internal_gross_salary,
                summary.internal_labor_cost,
            )
        ):
            summaries.append(summary)

    return summaries


def get_assignment_contribution_multiplier(
    env,
    multiplier_index: dict[str, float],
    employee_match: EmployeeMatch | None,
    employee_name: str,
) -> float:
    if employee_match and employee_match.employee_id:
        multiplier = multiplier_index.get(f"employee:{int(employee_match.employee_id)}")
        if multiplier:
            return multiplier
    for alias in employee_aliases(employee_name):
        multiplier = multiplier_index.get(f"alias:{alias}")
        if multiplier:
            return multiplier
    if env and employee_match and employee_match.employee_id:
        employee = env["hr.employee"].browse(employee_match.employee_id).exists()
        if employee and hasattr(employee, "_get_payroll_contribution_multiplier"):
            return employee._get_payroll_contribution_multiplier()
    return CCP_MULTIPLIER


def read_assignment_pair(sheet, row_idx: int) -> tuple[str, dict[int, float], dict[int, float]]:
    name = normalize_text(sheet.cell(row=row_idx, column=1).value)
    costs = {}
    hours = {}
    for col_idx, month in MONTH_BY_COLUMN.items():
        costs[month] = parse_amount(sheet.cell(row=row_idx, column=col_idx).value)
        hours[month] = parse_amount(sheet.cell(row=row_idx + 1, column=col_idx).value)
    return name, costs, hours


def parse_assignments(
    env,
    workbook_path: Path,
    projects_by_key: dict[str, ProjectPreview],
    unmatched: list[UnmatchedRow],
    warnings: list[MigrationWarning],
) -> tuple[list[AssignmentPreview], list[TimesheetPreview]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    employee_index = build_employee_index(env, workbook)
    employee_multiplier_index = load_employee_contribution_multipliers(workbook, employee_index)
    expected_hours = load_expected_hours(workbook)
    assignments = []
    timesheets = []

    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("summary_") or sheet_name in ASSIGNMENT_AGGREGATE_SHEETS:
            continue
        sheet = workbook[sheet_name]
        project_label = sheet_project_label(sheet_name)
        project = resolve_project_preview(projects_by_key, project_label)
        if not project:
            unmatched.append(UnmatchedRow("Assignments", sheet_name, "project", project_label, "Project sheet not found in 2026 project summary"))

        row_idx = 2
        while row_idx < sheet.max_row:
            row_type = fold_text(sheet.cell(row=row_idx, column=2).value)
            next_type = fold_text(sheet.cell(row=row_idx + 1, column=2).value)
            if "celkova cena prace" not in row_type or "odpracovane hodiny" not in next_type:
                row_idx += 1
                continue

            employee_name, costs, hours = read_assignment_pair(sheet, row_idx)
            if not employee_name or fold_text(employee_name) in {"total", "spolu"}:
                row_idx += 2
                continue
            employee_match = resolve_employee(employee_index, employee_name)
            if not employee_match:
                unmatched.append(UnmatchedRow("Assignments", f"{sheet_name}:{row_idx}", "employee", employee_name, "Employee not found"))
            contribution_multiplier = get_assignment_contribution_multiplier(
                env,
                employee_multiplier_index,
                employee_match,
                employee_name,
            )

            total_hours = round(sum(hours.values()), 2)
            total_ccp = round(sum(costs.values()), 2)
            if total_hours <= 0.0 and total_ccp <= 0.0:
                row_idx += 2
                continue
            if total_hours <= 0.0 and total_ccp > 0.0:
                warnings.append(MigrationWarning("Assignments", f"{sheet_name}:{row_idx}", "warning", "CCP exists with zero hours", employee_name))

            wage_ccp = round(total_ccp / total_hours, 4) if total_hours else 0.0
            wage_hm = round(wage_ccp / contribution_multiplier, 4) if wage_ccp and contribution_multiplier else 0.0
            monthly_ratios = {}
            active_months = []
            for month in range(1, 13):
                expected = expected_hours.get(month) or 0.0
                ratio = min(100.0, round((hours[month] / expected) * 100.0, 2)) if expected else 0.0
                monthly_ratios[month] = ratio
                if hours[month] or costs[month]:
                    active_months.append(month)
                    timesheets.append(TimesheetPreview(
                        sheet=sheet_name,
                        project_name=project.project_name if project else project_label,
                        employee_name=employee_name,
                        period=date(YEAR, month, 1),
                        hours_pp=round(hours[month], 2),
                        expected_hours=expected,
                        allocation_ratio=ratio,
                        ccp_amount=round(costs[month], 2),
                    ))

            if active_months:
                assignments.append(AssignmentPreview(
                    sheet=sheet_name,
                    project_name=project.project_name if project else project_label,
                    project_id=project.existing_project_id if project and project.existing_project_id else False,
                    employee_name=employee_name,
                    employee_id=employee_match.employee_id if employee_match else False,
                    employee_user_id=employee_match.user_id if employee_match else False,
                    employee_user_login=employee_match.user_login if employee_match else "",
                    employee_work_email=employee_match.work_email if employee_match else "",
                    employee_match_status=employee_match.match_status if employee_match else "missing",
                    employee_source_ref=employee_match.source_ref if employee_match else "",
                    total_hours=total_hours,
                    total_ccp=total_ccp,
                    wage_ccp=wage_ccp,
                    wage_hm=wage_hm,
                    contribution_multiplier=contribution_multiplier,
                    allocation_ratio=max(monthly_ratios.values() or [0.0]),
                    date_start=date(YEAR, min(active_months), 1),
                    date_end=project.date_end if project and project.date_end else date(YEAR, max(active_months), 1),
                    monthly_ratios=monthly_ratios,
                ))
            row_idx += 2

    return assignments, timesheets


def clean_apz_employee_label(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r".*->", "", text)
    text = re.sub(r"\bod\s+\d{1,2}[./]\d{1,2}[./]?\d{0,4}", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(do|od)\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(sestra)\b", "", text, flags=re.IGNORECASE)
    return normalize_text(text.strip(" -:>"))


def parse_apz_percent_entries(value: Any) -> list[tuple[str, float]]:
    text = normalize_text(value)
    if not text or "%" not in text:
        return []
    entries = []
    for match in re.finditer(r"(.+?)\s+(\d+(?:[,.]\d+)?)\s*%", text):
        employee_label = clean_apz_employee_label(match.group(1))
        if not employee_label:
            continue
        entries.append((employee_label, parse_amount(match.group(2))))
    return entries


def parse_apz_special_assignments(
    env,
    apz_workbook_path: Path,
    employee_workbook_path: Path,
    projects_by_key: dict[str, ProjectPreview],
    unmatched: list[UnmatchedRow],
    warnings: list[MigrationWarning],
) -> list[AssignmentPreview]:
    workbook = load_workbook(apz_workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    employee_workbook = load_workbook(employee_workbook_path, read_only=True, data_only=True)
    employee_index = build_employee_index(env, employee_workbook)
    apz_employee_index = build_summary_employee_short_index(employee_index, employee_workbook, "summary_APZ_2N")
    project = projects_by_key.get(project_key(APZ_PROJECT_NAME)) or projects_by_key.get(project_key("APZ 2N"))
    if not project:
        unmatched.append(UnmatchedRow("APZ Special Assignments", sheet.title, "project", APZ_PROJECT_NAME, "APZ project not found"))
        return []

    ratios_by_employee: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    matches_by_employee: dict[str, EmployeeMatch | None] = {}
    source_refs_by_employee: dict[str, set[str]] = defaultdict(set)

    for row_idx in range(3, sheet.max_row + 1):
        activity = normalize_text(sheet.cell(row=row_idx, column=1).value)
        position = normalize_text(sheet.cell(row=row_idx, column=3).value)
        regime = normalize_text(sheet.cell(row=row_idx, column=4).value)
        source_context = " / ".join(part for part in [activity, position, regime] if part)
        for col_idx, month in APZ_SPECIAL_MONTH_COLUMNS.items():
            for employee_label, ratio in parse_apz_percent_entries(sheet.cell(row=row_idx, column=col_idx).value):
                employee_match = None
                for key in employee_short_keys(employee_label):
                    employee_match = apz_employee_index.get(key)
                    if employee_match:
                        break
                if not employee_match:
                    employee_match = resolve_employee_short(employee_index, employee_label)
                employee_name = employee_match.name if employee_match else employee_label
                if employee_match is None:
                    unmatched.append(UnmatchedRow(
                        "APZ Special Assignments",
                        f"{sheet.title}:{row_idx}:{col_idx}",
                        "employee",
                        employee_label,
                        "Employee from APZ special workbook not found",
                        source_context,
                    ))
                matches_by_employee.setdefault(employee_name, employee_match)
                ratios_by_employee[employee_name][month] += ratio
                source_refs_by_employee[employee_name].add(f"{sheet.title}:{row_idx}:{col_idx}")

    assignments = []
    for employee_name, monthly_values in sorted(ratios_by_employee.items()):
        capped_ratios = {
            month: min(100.0, round(monthly_values.get(month, 0.0), 2))
            for month in range(1, 13)
        }
        active_months = [month for month, ratio in capped_ratios.items() if ratio]
        if not active_months:
            continue
        employee_match = matches_by_employee.get(employee_name)
        assignments.append(AssignmentPreview(
            sheet=f"apz_special:{sheet.title}",
            project_name=project.project_name,
            project_id=project.existing_project_id if project.existing_project_id else False,
            employee_name=employee_name,
            employee_id=employee_match.employee_id if employee_match else False,
            employee_user_id=employee_match.user_id if employee_match else False,
            employee_user_login=employee_match.user_login if employee_match else "",
            employee_work_email=employee_match.work_email if employee_match else "",
            employee_match_status=employee_match.match_status if employee_match else "missing",
            employee_source_ref=employee_match.source_ref if employee_match else "; ".join(sorted(source_refs_by_employee[employee_name])[:3]),
            total_hours=0.0,
            total_ccp=0.0,
            wage_ccp=0.0,
            wage_hm=0.0,
            contribution_multiplier=CCP_MULTIPLIER,
            allocation_ratio=max(capped_ratios.values() or [0.0]),
            date_start=date(YEAR, min(active_months), 1),
            date_end=project.date_end if project.date_end else date(YEAR, max(active_months), monthrange(YEAR, max(active_months))[1]),
            monthly_ratios=capped_ratios,
        ))

    warnings.append(MigrationWarning(
        "APZ Special Assignments",
        sheet.title,
        "info",
        "APZ special assignment workbook parsed as ratio-only assignment preview rows",
        str(apz_workbook_path),
    ))
    return assignments


def parse_cashflows(
    env,
    workbook_path: Path,
    sheet_name: str,
    projects_by_key: dict[str, ProjectPreview],
    unmatched: list[UnmatchedRow],
    warnings: list[MigrationWarning],
) -> list[CashflowPreview]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        if sheet_name == DEFAULT_CASHFLOW_SHEET and CASHFLOW_SHEET_FALLBACK in workbook.sheetnames:
            sheet_name = CASHFLOW_SHEET_FALLBACK
        else:
            raise ValueError(f"Cashflow sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    cashflows = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, max_col=36, values_only=True), start=3):
        project_label = normalize_text(row[20])
        fallback_label = normalize_text(row[1])
        label = project_label or fallback_label
        if not label:
            continue
        month_amounts = {
            month: parse_amount(row[col_idx - 1])
            for col_idx, month in CF_MONTH_COLUMNS.items()
        }
        receipt_amount = parse_amount(row[33])
        monthly_total = round(sum(month_amounts.values()), 2)
        if abs(monthly_total - round(receipt_amount, 2)) > 0.05:
            warnings.append(MigrationWarning("Cashflows", row_idx, "warning", "Monthly cashflow total differs from AH total", f"{label}: {monthly_total} != {receipt_amount}"))

        if is_non_project_cashflow(label, receipt_amount):
            handled_plan_row = (
                fold_text(label) in CF_PLAN_SUBTOTAL_LABELS
                or row_idx in CF_PLAN_EXPENSE_ROWS
                or (row_idx in CF_PLAN_INCOME_ROWS and not re.match(r"^\d+", label))
            )
            if not handled_plan_row:
                unmatched.append(UnmatchedRow("Cashflows", row_idx, "cashflow", label, "Non-project, negative, zero, or balance cashflow row", str(receipt_amount)))
            continue
        negative_months = {
            month: amount
            for month, amount in month_amounts.items()
            if amount < 0.0
        }
        if negative_months:
            unmatched.append(UnmatchedRow(
                "Cashflows",
                row_idx,
                "cashflow",
                label,
                "Project cashflow row contains negative monthly value",
                ", ".join(f"{month}: {amount}" for month, amount in sorted(negative_months.items())),
            ))
            continue

        project = resolve_project_preview(projects_by_key, label)
        if not project:
            unmatched.append(UnmatchedRow("Cashflows", row_idx, "project", label, "Cashflow project not found in 2026 project summary", str(receipt_amount)))
            continue

        existing_project = env["tenenet.project"].with_context(active_test=False).browse(project.existing_project_id).exists() if env and project.existing_project_id else None
        cashflows.append(CashflowPreview(
            source_row=row_idx,
            source_sheet=sheet_name,
            program_label=normalize_text(row[0]),
            project_label=label,
            project_id=existing_project.id if existing_project else False,
            receipt_date=date(YEAR, 1, 1),
            receipt_amount=receipt_amount,
            receipt_note=normalize_text(row[34]) or normalize_text(row[15]),
            month_amounts=month_amounts,
        ))

    return cashflows


def parse_cashflow_plan_rows(
    env,
    workbook_path: Path,
    sheet_name: str,
    projects_by_key: dict[str, ProjectPreview],
    warnings: list[MigrationWarning],
) -> list[CashflowPlanPreview]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        if sheet_name == DEFAULT_CASHFLOW_SHEET and CASHFLOW_SHEET_FALLBACK in workbook.sheetnames:
            sheet_name = CASHFLOW_SHEET_FALLBACK
        else:
            raise ValueError(f"Cashflow sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = []

    for row_idx in list(CF_PLAN_INCOME_ROWS) + list(CF_PLAN_EXPENSE_ROWS):
        row = [sheet.cell(row_idx, column).value for column in range(1, 36)]
        right_label = normalize_text(row[20])
        fallback_label = normalize_text(row[1])
        label = right_label or fallback_label
        if not label:
            continue
        if fold_text(label) in CF_PLAN_SUBTOTAL_LABELS:
            continue

        row_type = cashflow_plan_row_type(row_idx, label)
        month_amounts = {
            month: parse_amount(row[col_idx - 1])
            for col_idx, month in CF_MONTH_COLUMNS.items()
        }
        if not any(month_amounts.values()):
            continue

        if row_type == "income":
            total_amount = parse_amount(row[33])
            project = resolve_project_preview(projects_by_key, label)
            if project and total_amount > 0.0:
                continue
            if total_amount > 0.0 and re.match(r"^\d+", label):
                continue

        monthly_total = round(sum(month_amounts.values()), 2)
        total_amount = parse_amount(row[33])
        if abs(monthly_total - round(total_amount, 2)) > 0.05:
            warnings.append(MigrationWarning(
                "Cashflow plan",
                row_idx,
                "warning",
                "Monthly cashflow plan total differs from AH total",
                f"{label}: {monthly_total} != {total_amount}",
            ))

        row_key = cashflow_plan_row_key(row_type, label)
        rows.append(CashflowPlanPreview(
            source_row=row_idx,
            source_sheet=sheet_name,
            row_key=row_key,
            row_label=label,
            row_type=row_type,
            section_label="Príjmy" if row_type == "income" else "Výdavky",
            program_label=normalize_text(row[0]),
            sequence=100 + row_idx,
            actual_mapping_key=row_key,
            month_amounts=month_amounts,
        ))

    return rows


def as_date_text(value: date | None) -> str:
    return value.isoformat() if value else ""


def month_end(value: date | None) -> date | None:
    if not value:
        return None
    return date(value.year, value.month, monthrange(value.year, value.month)[1])


def months_in_year_range(start: date | None, end: date | None, year: int = YEAR) -> list[int]:
    range_start = start or date(year, 1, 1)
    range_end = end or date(year, 12, 31)
    first = max(date(year, 1, 1), date(range_start.year, range_start.month, 1))
    last = min(date(year, 12, 31), date(range_end.year, range_end.month, 1))
    if first > last:
        return []
    months = []
    current = first
    while current <= last:
        months.append(current.month)
        if current.month == 12:
            break
        current = date(current.year, current.month + 1, 1)
    return months


def build_employee_match_report(assignments: list[AssignmentPreview]) -> list[list[Any]]:
    grouped = {}
    for assignment in assignments:
        key = assignment.employee_name
        row = grouped.setdefault(key, {
            "employee_name": assignment.employee_name,
            "employee_id": assignment.employee_id,
            "employee_user_id": assignment.employee_user_id,
            "employee_user_login": assignment.employee_user_login,
            "employee_work_email": assignment.employee_work_email,
            "employee_match_status": assignment.employee_match_status,
            "employee_source_ref": assignment.employee_source_ref,
            "assignment_count": 0,
            "project_names": set(),
            "sheets": set(),
        })
        row["assignment_count"] += 1
        row["project_names"].add(assignment.project_name)
        row["sheets"].add(assignment.sheet)
        if not row["employee_id"] and assignment.employee_id:
            row["employee_id"] = assignment.employee_id
        if not row["employee_user_id"] and assignment.employee_user_id:
            row["employee_user_id"] = assignment.employee_user_id
            row["employee_user_login"] = assignment.employee_user_login
        if not row["employee_work_email"] and assignment.employee_work_email:
            row["employee_work_email"] = assignment.employee_work_email
    return [[
        row["employee_name"],
        row["employee_id"],
        row["employee_user_id"],
        row["employee_user_login"],
        row["employee_work_email"],
        row["employee_match_status"],
        row["employee_source_ref"],
        row["assignment_count"],
        "; ".join(sorted(row["project_names"])),
        "; ".join(sorted(row["sheets"])),
    ] for row in sorted(grouped.values(), key=lambda item: fold_text(item["employee_name"]))]


def write_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(normalize_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)


def write_preview_workbook(report_path: Path, result: dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    projects = result["projects"]
    assignments = result["assignments"]
    timesheets = result["timesheets"]
    cashflows = result["cashflows"]
    cashflow_plan_rows = result.get("cashflow_plan_rows", [])
    employee_monthly_summaries = result.get("employee_monthly_summaries", [])
    unmatched = result["unmatched"]
    warnings = result["warnings"]
    import_actions = result.get("import_actions", [])

    write_sheet(workbook, "Summary", ["Metric", "Value"], [
        ["Projects", len(projects)],
        ["Assignments", len(assignments)],
        ["Timesheets", len(timesheets)],
        ["Cashflows", len(cashflows)],
        ["Cashflow plan rows", len(cashflow_plan_rows)],
        ["Employee monthly summaries", len(employee_monthly_summaries)],
        ["Unmatched", len(unmatched)],
        ["Warnings", len(warnings)],
        ["Import actions", len(import_actions)],
    ])
    if import_actions:
        action_counter = Counter(f"{item.model}:{item.action}:{item.status}" for item in import_actions)
        write_sheet(workbook, "Import Summary", ["Metric", "Value"], [
            [key, value] for key, value in sorted(action_counter.items())
        ])
    write_sheet(workbook, "Projects", [
        "source_row", "project_name", "description", "contract_number", "recipient_text", "recipient_id",
        "donor_text", "donor_id", "project_type_raw", "project_type", "program_raw", "program_code",
        "program_id", "budget_amount", "received_2026", "date_start", "date_end", "existing_project_id", "match_status",
    ], [[
        item.source_row, item.project_name, item.description, item.contract_number, item.recipient_text, item.recipient_id,
        item.donor_text, item.donor_id, item.project_type_raw, item.project_type, item.program_raw, item.program_code,
        item.program_id, item.budget_amount, item.received_2026, as_date_text(item.date_start), as_date_text(item.date_end),
        item.existing_project_id, item.match_status,
    ] for item in projects])
    write_sheet(workbook, "Assignments", [
        "sheet", "project_name", "project_id", "employee_name", "employee_id", "employee_user_id", "employee_user_login",
        "employee_work_email", "employee_match_status", "employee_source_ref", "total_hours", "total_ccp",
        "wage_ccp", "wage_hm", "contribution_multiplier", "allocation_ratio", "date_start", "date_end",
        "ratio_01", "ratio_02", "ratio_03", "ratio_04", "ratio_05", "ratio_06",
        "ratio_07", "ratio_08", "ratio_09", "ratio_10", "ratio_11", "ratio_12",
    ], [[
        item.sheet, item.project_name, item.project_id, item.employee_name, item.employee_id, item.employee_user_id,
        item.employee_user_login, item.employee_work_email, item.employee_match_status, item.employee_source_ref,
        item.total_hours, item.total_ccp,
        item.wage_ccp, item.wage_hm, item.contribution_multiplier, item.allocation_ratio, as_date_text(item.date_start), as_date_text(item.date_end),
        *[item.monthly_ratios.get(month, 0.0) for month in range(1, 13)],
    ] for item in assignments])
    write_sheet(workbook, "Employee Matches", [
        "employee_name", "employee_id", "employee_user_id", "employee_user_login", "employee_work_email",
        "employee_match_status", "employee_source_ref", "assignment_count", "project_names", "sheets",
    ], build_employee_match_report(assignments))
    write_sheet(workbook, "Employee Monthly Summaries", [
        "sheet", "employee_name", "employee_id", "employee_match_status", "employee_source_ref",
        "capacity_incl_01", "capacity_incl_02", "capacity_incl_03", "capacity_incl_04", "capacity_incl_05", "capacity_incl_06",
        "capacity_incl_07", "capacity_incl_08", "capacity_incl_09", "capacity_incl_10", "capacity_incl_11", "capacity_incl_12",
        "capacity_01", "capacity_02", "capacity_03", "capacity_04", "capacity_05", "capacity_06",
        "capacity_07", "capacity_08", "capacity_09", "capacity_10", "capacity_11", "capacity_12",
        "gross_01", "gross_02", "gross_03", "gross_04", "gross_05", "gross_06",
        "gross_07", "gross_08", "gross_09", "gross_10", "gross_11", "gross_12",
        "ccp_01", "ccp_02", "ccp_03", "ccp_04", "ccp_05", "ccp_06",
        "ccp_07", "ccp_08", "ccp_09", "ccp_10", "ccp_11", "ccp_12",
        "worked_01", "worked_02", "worked_03", "worked_04", "worked_05", "worked_06",
        "worked_07", "worked_08", "worked_09", "worked_10", "worked_11", "worked_12",
        "holidays_01", "holidays_02", "holidays_03", "holidays_04", "holidays_05", "holidays_06",
        "holidays_07", "holidays_08", "holidays_09", "holidays_10", "holidays_11", "holidays_12",
        "vacation_01", "vacation_02", "vacation_03", "vacation_04", "vacation_05", "vacation_06",
        "vacation_07", "vacation_08", "vacation_09", "vacation_10", "vacation_11", "vacation_12",
        "doctor_01", "doctor_02", "doctor_03", "doctor_04", "doctor_05", "doctor_06",
        "doctor_07", "doctor_08", "doctor_09", "doctor_10", "doctor_11", "doctor_12",
        "internal_gross_01", "internal_gross_02", "internal_gross_03", "internal_gross_04", "internal_gross_05", "internal_gross_06",
        "internal_gross_07", "internal_gross_08", "internal_gross_09", "internal_gross_10", "internal_gross_11", "internal_gross_12",
        "internal_ccp_01", "internal_ccp_02", "internal_ccp_03", "internal_ccp_04", "internal_ccp_05", "internal_ccp_06",
        "internal_ccp_07", "internal_ccp_08", "internal_ccp_09", "internal_ccp_10", "internal_ccp_11", "internal_ccp_12",
    ], [[
        item.sheet, item.employee_name, item.employee_id, item.employee_match_status, item.employee_source_ref,
        *[item.capacity_hours_incl.get(month, 0.0) for month in range(1, 13)],
        *[item.capacity_hours.get(month, 0.0) for month in range(1, 13)],
        *[item.total_gross_salary.get(month, 0.0) for month in range(1, 13)],
        *[item.total_labor_cost.get(month, 0.0) for month in range(1, 13)],
        *[item.worked_hours.get(month, 0.0) for month in range(1, 13)],
        *[item.holidays_hours.get(month, 0.0) for month in range(1, 13)],
        *[item.vacation_hours.get(month, 0.0) for month in range(1, 13)],
        *[item.doctor_hours.get(month, 0.0) for month in range(1, 13)],
        *[item.internal_gross_salary.get(month, 0.0) for month in range(1, 13)],
        *[item.internal_labor_cost.get(month, 0.0) for month in range(1, 13)],
    ] for item in employee_monthly_summaries])
    write_sheet(workbook, "Timesheets", [
        "sheet", "project_name", "employee_name", "period", "hours_pp", "expected_hours", "allocation_ratio", "ccp_amount",
    ], [[
        item.sheet, item.project_name, item.employee_name, as_date_text(item.period), item.hours_pp, item.expected_hours,
        item.allocation_ratio, item.ccp_amount,
    ] for item in timesheets])
    write_sheet(workbook, "Cashflows", [
        "source_sheet", "source_row", "program_label", "project_label", "project_id", "receipt_date", "receipt_amount", "receipt_note",
        "month_01", "month_02", "month_03", "month_04", "month_05", "month_06",
        "month_07", "month_08", "month_09", "month_10", "month_11", "month_12",
    ], [[
        item.source_sheet, item.source_row, item.program_label, item.project_label, item.project_id, as_date_text(item.receipt_date),
        item.receipt_amount, item.receipt_note,
        *[item.month_amounts.get(month, 0.0) for month in range(1, 13)],
    ] for item in cashflows])
    write_sheet(workbook, "Cashflow Plan", [
        "source_sheet", "source_row", "row_key", "row_label", "row_type", "section_label", "program_label",
        "sequence", "actual_mapping_key",
        "month_01", "month_02", "month_03", "month_04", "month_05", "month_06",
        "month_07", "month_08", "month_09", "month_10", "month_11", "month_12",
    ], [[
        item.source_sheet, item.source_row, item.row_key, item.row_label, item.row_type, item.section_label,
        item.program_label, item.sequence, item.actual_mapping_key,
        *[item.month_amounts.get(month, 0.0) for month in range(1, 13)],
    ] for item in cashflow_plan_rows])
    write_sheet(workbook, "Unmatched", ["source", "row", "kind", "key", "reason", "context"], [
        [item.source, item.row, item.kind, item.key, item.reason, item.context]
        for item in unmatched
    ])
    write_sheet(workbook, "Warnings", ["source", "row", "level", "message", "raw_value"], [
        [item.source, item.row, item.level, item.message, item.raw_value]
        for item in warnings
    ])
    if import_actions:
        write_sheet(workbook, "Import Actions", ["source", "action", "model", "record_id", "key", "status", "message"], [
            [item.source, item.action, item.model, item.record_id, item.key, item.status, item.message]
            for item in import_actions
        ])

    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)


def build_migration_preview(
    env,
    project_summary_path,
    assignments_path,
    cashflow_path,
    report_path,
    cashflow_sheet=DEFAULT_CASHFLOW_SHEET,
    apz_assignments_path=None,
    project_program_bindings_path=None,
) -> dict[str, Any]:
    project_summary_path = Path(project_summary_path)
    assignments_path = Path(assignments_path)
    cashflow_path = Path(cashflow_path)
    report_path = Path(report_path)
    project_program_bindings_path = Path(project_program_bindings_path) if project_program_bindings_path else assignments_path

    warnings: list[MigrationWarning] = []
    unmatched: list[UnmatchedRow] = []

    projects = parse_project_summary(env, project_summary_path, warnings)
    append_assignment_project_previews(env, assignments_path, project_program_bindings_path, projects, warnings)
    projects_by_key = build_project_index(projects)
    assignments, timesheets = parse_assignments(env, assignments_path, projects_by_key, unmatched, warnings)
    employee_monthly_summaries = parse_employee_monthly_summaries(env, assignments_path, unmatched)
    if apz_assignments_path:
        assignments.extend(parse_apz_special_assignments(
            env,
            Path(apz_assignments_path),
            assignments_path,
            projects_by_key,
            unmatched,
            warnings,
        ))
    cashflows = parse_cashflows(env, cashflow_path, cashflow_sheet, projects_by_key, unmatched, warnings)
    cashflow_plan_rows = parse_cashflow_plan_rows(env, cashflow_path, cashflow_sheet, projects_by_key, warnings)

    result = {
        "projects": projects,
        "assignments": assignments,
        "timesheets": timesheets,
        "employee_monthly_summaries": employee_monthly_summaries,
        "cashflows": cashflows,
        "cashflow_plan_rows": cashflow_plan_rows,
        "unmatched": unmatched,
        "warnings": warnings,
        "summary": Counter({
            "projects": len(projects),
            "assignments": len(assignments),
            "timesheets": len(timesheets),
            "employee_monthly_summaries": len(employee_monthly_summaries),
            "cashflows": len(cashflows),
            "cashflow_plan_rows": len(cashflow_plan_rows),
            "unmatched": len(unmatched),
            "warnings": len(warnings),
        }),
    }
    write_preview_workbook(report_path, result)
    return result


def project_values(env, project: ProjectPreview) -> dict[str, Any]:
    values = {
        "name": project.project_name,
        "description": project.description,
        "contract_number": project.contract_number,
        "project_type": project.project_type,
        "date_start": project.date_start,
        "date_end": project.date_end,
        "active": True,
    }
    if project.recipient_id:
        values["recipient_partner_id"] = project.recipient_id
    if project.donor_id:
        values["donor_id"] = project.donor_id
    admin_program = env["tenenet.program"].with_context(active_test=False).search([("code", "=", "ADMIN_TENENET")], limit=1)
    program_ids = []
    if project.project_type != "medzinarodny" and project.program_id:
        program_ids.append(project.program_id)
    if admin_program:
        program_ids.append(admin_program.id)
    if program_ids:
        values["program_ids"] = [(6, 0, list(dict.fromkeys(program_ids)))]
    return {key: value for key, value in values.items() if value not in (None, "")}


def build_import_project_maps(env, projects: list[ProjectPreview]) -> tuple[dict[str, Any], dict[str, Any]]:
    Project = env["tenenet.project"].sudo().with_context(active_test=False)
    records_by_name = {}
    records_by_key = {}
    for project in projects:
        record = Project.browse(project.existing_project_id).exists() if project.existing_project_id else Project.search(
            [("name", "=", project.project_name)],
            limit=1,
        )
        if record:
            records_by_name[project.project_name] = record
            for value in (project.project_name, project.description, project.contract_number):
                key = project_key(value)
                if key:
                    records_by_key.setdefault(key, record)
    return records_by_name, records_by_key


def apply_projects(env, projects: list[ProjectPreview], actions: list[ImportAction]) -> dict[str, Any]:
    Project = env["tenenet.project"].sudo().with_context(active_test=False)
    records_by_name = {}
    for project in projects:
        record = Project.browse(project.existing_project_id).exists() if project.existing_project_id else Project.search(
            [("name", "=", project.project_name)],
            limit=1,
        )
        values = project_values(env, project)
        if record:
            record.write(values)
            action = "update"
        else:
            record = Project.create(values)
            action = "create"
        records_by_name[project.project_name] = record
        actions.append(ImportAction("Projects", action, "tenenet.project", record.id, project.project_name, "ok"))
    return records_by_name


def assignment_values(assignment: AssignmentPreview, project_record) -> dict[str, Any]:
    assignment_end = project_record.date_end or month_end(assignment.date_end)
    values = {
        "employee_id": assignment.employee_id,
        "project_id": project_record.id,
        "date_start": assignment.date_start,
        "date_end": assignment_end,
        "allocation_ratio": max(0.01, min(100.0, assignment.allocation_ratio or 0.0)),
        "wage_hm": assignment.wage_hm or 0.0,
        "active": True,
    }
    program = project_record._get_effective_reporting_program() or project_record.program_ids[:1]
    if program:
        values["program_id"] = program.id
    return values


def apply_employee_multipliers(env, assignments: list[AssignmentPreview], actions: list[ImportAction]) -> None:
    Employee = env["hr.employee"].sudo().with_context(active_test=False)
    multipliers_by_employee: dict[int, float] = {}
    for assignment in assignments:
        if assignment.employee_id and assignment.contribution_multiplier:
            multipliers_by_employee[int(assignment.employee_id)] = assignment.contribution_multiplier

    for employee_id, multiplier in sorted(multipliers_by_employee.items()):
        employee = Employee.browse(employee_id).exists()
        if not employee:
            continue
        current = employee._get_payroll_contribution_multiplier() if hasattr(employee, "_get_payroll_contribution_multiplier") else CCP_MULTIPLIER
        if math.isclose(current, multiplier, rel_tol=0.0, abs_tol=0.0001):
            continue
        employee.write({"tenenet_payroll_contribution_multiplier": multiplier})
        actions.append(ImportAction(
            "Employees",
            "update",
            "hr.employee",
            employee.id,
            employee.display_name,
            "ok",
            f"tenenet_payroll_contribution_multiplier={multiplier}",
        ))


def apply_employee_monthly_summaries(env, summaries: list[EmployeeMonthlySummaryPreview], actions: list[ImportAction]) -> None:
    Cost = env["tenenet.employee.tenenet.cost"].sudo().with_context(active_test=False)
    for summary in summaries:
        if not summary.employee_id:
            actions.append(ImportAction(
                "Employee monthly summary",
                "skip",
                "tenenet.employee.tenenet.cost",
                False,
                summary.employee_name,
                "skipped",
                "Employee not matched",
            ))
            continue

        touched = 0
        for month in range(1, 13):
            period = date(YEAR, month, 1)
            record = Cost.search([
                ("employee_id", "=", summary.employee_id),
                ("period", "=", period),
            ], limit=1)
            if not record:
                record = Cost.create({
                    "employee_id": summary.employee_id,
                    "period": period,
                })
            values = {
                "imported_from_migration_workbook": True,
                "imported_capacity_hours_incl": summary.capacity_hours_incl.get(month, 0.0),
                "imported_capacity_hours": summary.capacity_hours.get(month, 0.0),
                "imported_total_gross_salary": summary.total_gross_salary.get(month, 0.0),
                "imported_total_labor_cost": summary.total_labor_cost.get(month, 0.0),
                "imported_worked_hours": summary.worked_hours.get(month, 0.0),
                "imported_holidays_hours": summary.holidays_hours.get(month, 0.0),
                "imported_vacation_hours": summary.vacation_hours.get(month, 0.0),
                "imported_doctor_hours": summary.doctor_hours.get(month, 0.0),
                "imported_internal_gross_salary": summary.internal_gross_salary.get(month, 0.0),
                "imported_internal_labor_cost": summary.internal_labor_cost.get(month, 0.0),
            }
            if any(record[field_name] != value for field_name, value in values.items()):
                record.write(values)
                touched += 1

        actions.append(ImportAction(
            "Employee monthly summary",
            "upsert",
            "tenenet.employee.tenenet.cost",
            False,
            summary.employee_name,
            "ok",
            f"{touched} monthly rows updated",
        ))


def find_existing_assignment(env, assignment: AssignmentPreview, project_record):
    Assignment = env["tenenet.project.assignment"].sudo().with_context(active_test=False)
    desired_end = project_record.date_end or month_end(assignment.date_end)
    candidates = Assignment.search([
        ("project_id", "=", project_record.id),
        ("employee_id", "=", assignment.employee_id),
        ("date_start", "=", assignment.date_start),
    ])
    if not candidates:
        return Assignment.browse()
    exact_end = candidates.filtered(lambda rec: rec.date_end == desired_end)
    if exact_end:
        candidates = exact_end
    wage_hm = assignment.wage_hm or 0.0
    exact = candidates.filtered(lambda rec: abs((rec.wage_hm or 0.0) - wage_hm) < 0.0001)
    if exact:
        return exact[:1]
    ratio = assignment.allocation_ratio or 0.0
    exact = candidates.filtered(lambda rec: abs((rec.allocation_ratio or 0.0) - ratio) < 0.01)
    return exact[:1] if exact else candidates[:1]


def timesheets_by_assignment_source(timesheets: list[TimesheetPreview]) -> dict[tuple[str, str, str], dict[date, TimesheetPreview]]:
    grouped: dict[tuple[str, str, str], dict[date, TimesheetPreview]] = defaultdict(dict)
    for timesheet in timesheets:
        grouped[(timesheet.sheet, timesheet.project_name, timesheet.employee_name)][timesheet.period] = timesheet
    return grouped


def set_assignment_month_ratios(assignment_record, preview: AssignmentPreview) -> None:
    scoped_months = months_in_year_range(assignment_record.date_start, assignment_record.date_end)
    ratios = {
        month: max(0.0, min(100.0, round(float(preview.monthly_ratios.get(month, 0.0) or 0.0), 2)))
        for month in (scoped_months or range(1, 13))
    }
    if ratios:
        assignment_record.set_month_ratios(YEAR, ratios)


def apply_timesheets(env, assignment_record, preview: AssignmentPreview, source_timesheets: dict[date, TimesheetPreview], actions: list[ImportAction]) -> None:
    Timesheet = env["tenenet.project.timesheet"].sudo()
    for period, timesheet in sorted(source_timesheets.items()):
        sheet = Timesheet._get_or_create_for_assignment_period(assignment_record, period)
        sheet.write({
            "hours_pp": timesheet.hours_pp,
            "labor_cost_override": timesheet.ccp_amount,
        })
        actions.append(ImportAction(
            "Timesheets",
            "upsert",
            "tenenet.project.timesheet",
            sheet.id,
            f"{preview.project_name} / {preview.employee_name} / {period.isoformat()}",
            "ok",
            f"hours_pp={timesheet.hours_pp}; labor_cost_override={timesheet.ccp_amount}",
        ))


def apply_assignments(env, result: dict[str, Any], project_records: dict[str, Any], actions: list[ImportAction]) -> dict[tuple[str, int], Any]:
    Assignment = env["tenenet.project.assignment"].sudo().with_context(
        skip_tenenet_assignment_capacity_check=True,
        active_test=False,
    )
    source_timesheets = timesheets_by_assignment_source(result["timesheets"])
    primary_by_project_employee: dict[tuple[str, int], Any] = {}

    for preview in result["assignments"]:
        key_text = f"{preview.project_name} / {preview.employee_name} / {preview.sheet}"
        if not preview.employee_id:
            actions.append(ImportAction("Assignments", "skip", "tenenet.project.assignment", False, key_text, "skipped", "Employee not matched"))
            continue
        project_record = project_records.get(preview.project_name)
        if not project_record:
            actions.append(ImportAction("Assignments", "skip", "tenenet.project.assignment", False, key_text, "skipped", "Project not available"))
            continue

        project_employee_key = (preview.project_name, int(preview.employee_id))
        is_ratio_only = not preview.total_hours and not preview.total_ccp
        if is_ratio_only and project_employee_key in primary_by_project_employee:
            assignment_record = primary_by_project_employee[project_employee_key].with_context(skip_tenenet_assignment_capacity_check=True)
            set_assignment_month_ratios(assignment_record, preview)
            actions.append(ImportAction("Assignments", "update_ratios", "tenenet.project.assignment", assignment_record.id, key_text, "ok", "Ratio-only row merged into existing assignment"))
            continue

        existing = find_existing_assignment(env, preview, project_record)
        values = assignment_values(preview, project_record)
        if existing:
            assignment_record = existing.with_context(skip_tenenet_assignment_capacity_check=True)
            assignment_record.write(values)
            action = "update"
        else:
            assignment_record = Assignment.create(values)
            action = "create"

        if not is_ratio_only:
            primary_by_project_employee.setdefault(project_employee_key, assignment_record)
        set_assignment_month_ratios(assignment_record, preview)
        actions.append(ImportAction("Assignments", action, "tenenet.project.assignment", assignment_record.id, key_text, "ok"))

        source_key = (preview.sheet, preview.project_name, preview.employee_name)
        apply_timesheets(env, assignment_record, preview, source_timesheets.get(source_key, {}), actions)

    return primary_by_project_employee


def cashflow_month_amounts_for_receipt(month_amounts: dict[int, float], receipt_amount: float, currency) -> tuple[dict[int, float], float]:
    rounded = {
        month: currency.round(float(amount or 0.0))
        for month, amount in month_amounts.items()
        if abs(amount or 0.0) > 0.00001
    }
    if not rounded:
        return rounded, 0.0

    receipt_total = currency.round(float(receipt_amount or 0.0))
    monthly_total = currency.round(sum(rounded.values()))
    excess = currency.round(monthly_total - receipt_total)
    if excess <= 0.0:
        return rounded, 0.0

    # Imported Excel rows sometimes exceed AH by a few cents after per-month rounding.
    if excess <= currency.rounding * 10:
        last_month = max(month for month, amount in rounded.items() if amount > 0.0)
        rounded[last_month] = currency.round(rounded[last_month] - excess)
        if abs(rounded[last_month]) < 0.00001:
            rounded.pop(last_month)
        return rounded, -excess

    return rounded, 0.0


def receipt_needs_write(receipt, values: dict[str, Any]) -> bool:
    for field_name, value in values.items():
        current = receipt[field_name]
        if hasattr(current, "ids"):
            current = current.id if current else False
        if current != value:
            return True
    return False


def cashflow_month_amounts_match(receipt, month_amounts: dict[int, float], currency) -> bool:
    existing = {
        cashflow.month: currency.round(cashflow.amount or 0.0)
        for cashflow in receipt.cashflow_ids
    }
    target = {
        month: currency.round(amount or 0.0)
        for month, amount in month_amounts.items()
        if abs(amount or 0.0) > 0.00001
    }
    return existing == target


def apply_cashflows(env, result: dict[str, Any], project_records: dict[str, Any], actions: list[ImportAction]) -> None:
    Receipt = env["tenenet.project.receipt"].sudo().with_context(skip_tenenet_receipt_auto_cashflow=True)
    projects_by_key = build_project_index(result["projects"])
    for cashflow in result["cashflows"]:
        project_preview = resolve_project_preview(projects_by_key, cashflow.project_label)
        project_record = project_records.get(project_preview.project_name) if project_preview else None
        key_text = f"{cashflow.source_sheet} row {cashflow.source_row}: {cashflow.project_label}"
        if not project_record:
            actions.append(ImportAction("Cashflows", "skip", "tenenet.project.receipt", False, key_text, "skipped", "Project not available"))
            continue
        source_label = cashflow.receipt_note or cashflow.project_label
        marker = f"Import z hárka {cashflow.source_sheet}, riadok {cashflow.source_row}: {source_label}"
        legacy_marker = f"MIGRATION CF 2026 row {cashflow.source_row}: {source_label}"
        receipt = Receipt.search([
            ("project_id", "=", project_record.id),
            "|",
            ("note", "=", marker),
            ("note", "=", legacy_marker),
        ], limit=1)
        values = {
            "project_id": project_record.id,
            "date_received": cashflow.receipt_date,
            "amount": cashflow.receipt_amount,
            "note": marker,
        }
        if receipt:
            if receipt_needs_write(receipt, values):
                receipt.write(values)
                action = "update"
            else:
                action = "noop"
        else:
            receipt = Receipt.create(values)
            action = "create"
        currency = receipt.currency_id or env.company.currency_id
        month_amounts, rounding_adjustment = cashflow_month_amounts_for_receipt(
            cashflow.month_amounts,
            cashflow.receipt_amount,
            currency,
        )
        if not cashflow_month_amounts_match(receipt, month_amounts, currency):
            receipt.set_cashflow_month_amounts(YEAR, month_amounts)
            if action == "noop":
                action = "update"
        detail = f"{len(receipt.cashflow_ids)} monthly rows"
        if rounding_adjustment:
            detail = f"{detail}; rounding adjustment {rounding_adjustment:+.2f}"
        actions.append(ImportAction("Cashflows", action, "tenenet.project.receipt", receipt.id, key_text, "ok", detail))


def apply_cashflow_plan_rows(env, result: dict[str, Any], actions: list[ImportAction]) -> None:
    Override = env["tenenet.cashflow.global.override"].sudo()
    for plan_row in result.get("cashflow_plan_rows", []):
        key_text = f"{plan_row.source_sheet} row {plan_row.source_row}: {plan_row.row_label}"
        action = "noop"
        for month in range(1, 13):
            period = date(YEAR, month, 1)
            values = {
                "period": period,
                "row_key": plan_row.row_key,
                "row_label": plan_row.row_label,
                "row_type": plan_row.row_type,
                "section_label": plan_row.section_label,
                "program_label": plan_row.program_label,
                "project_label": plan_row.row_label,
                "sequence": plan_row.sequence,
                "amount": plan_row.month_amounts.get(month, 0.0),
                "currency_id": env.company.currency_id.id,
                "source_kind": "workbook",
                "source_sheet": plan_row.source_sheet,
                "source_row": plan_row.source_row,
                "actual_mapping_key": plan_row.actual_mapping_key,
            }
            existing = Override.search([
                ("period", "=", period),
                ("row_key", "=", plan_row.row_key),
            ], limit=1)
            if existing:
                existing.with_context(
                    _cashflow_override_adjusting=True,
                    _skip_finance_monthly_comparison_sync=True,
                ).write(values)
                action = "update"
            else:
                Override.with_context(
                    _cashflow_override_adjusting=True,
                    _skip_finance_monthly_comparison_sync=True,
                ).create(values)
                if action == "noop":
                    action = "create"
        actions.append(ImportAction(
            "Cashflow plan",
            action,
            "tenenet.cashflow.global.override",
            False,
            key_text,
            "ok",
            "12 monthly rows",
        ))


def apply_migration_preview(env, result: dict[str, Any]) -> dict[str, Any]:
    actions: list[ImportAction] = []
    project_records = apply_projects(env, result["projects"], actions)
    apply_employee_multipliers(env, result["assignments"], actions)
    apply_employee_monthly_summaries(env, result.get("employee_monthly_summaries", []), actions)
    apply_assignments(env, result, project_records, actions)
    apply_cashflows(env, result, project_records, actions)
    apply_cashflow_plan_rows(env, result, actions)
    result["import_actions"] = actions
    result["summary"].update({
        "import_actions": len(actions),
        "import_action_errors": len([action for action in actions if action.status == "error"]),
        "import_action_skipped": len([action for action in actions if action.status == "skipped"]),
    })
    return result


def import_migration_data(
    env,
    project_summary_path,
    assignments_path,
    cashflow_path,
    report_path,
    cashflow_sheet=DEFAULT_CASHFLOW_SHEET,
    apz_assignments_path=None,
    project_program_bindings_path=None,
) -> dict[str, Any]:
    result = build_migration_preview(
        env,
        project_summary_path,
        assignments_path,
        cashflow_path,
        report_path,
        cashflow_sheet=cashflow_sheet,
        apz_assignments_path=apz_assignments_path,
        project_program_bindings_path=project_program_bindings_path,
    )
    apply_migration_preview(env, result)
    write_preview_workbook(Path(report_path), result)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Build dry-run XLSX preview or import TENENET project migration workbooks.")
    parser.add_argument("--project-summary", required=True, type=Path)
    parser.add_argument("--assignments", required=True, type=Path)
    parser.add_argument("--cashflow", required=True, type=Path)
    parser.add_argument("--cashflow-sheet", default=DEFAULT_CASHFLOW_SHEET)
    parser.add_argument("--apz-assignments", type=Path)
    parser.add_argument("--project-program-bindings", type=Path)
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--apply", action="store_true", help="Create/update Odoo records. Default is dry-run only.")
    args = parser.parse_args()

    current_env = globals().get("env")
    if current_env is None:
        raise SystemExit(
            "Odoo env is not available. Run this script from `odoo-bin shell` and pass "
            "`init_globals={'env': env}` to runpy.run_path()."
        )

    runner = import_migration_data if args.apply else build_migration_preview
    result = runner(
        current_env,
        args.project_summary,
        args.assignments,
        args.cashflow,
        args.report,
        cashflow_sheet=args.cashflow_sheet,
        apz_assignments_path=args.apz_assignments,
        project_program_bindings_path=args.project_program_bindings,
    )
    action_count = len(result.get("import_actions", []))
    print(
        "%s written to %s: %s projects, %s assignments, %s timesheets, %s cashflows, %s unmatched, %s warnings, %s import actions"
        % (
            "Import report" if args.apply else "Preview",
            args.report,
            len(result["projects"]),
            len(result["assignments"]),
            len(result["timesheets"]),
            len(result["cashflows"]),
            len(result["unmatched"]),
            len(result["warnings"]),
            action_count,
        )
    )


if __name__ == "__main__":
    main()
