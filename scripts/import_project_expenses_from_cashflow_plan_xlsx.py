#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import unicodedata
from calendar import monthrange
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from odoo import Command


DEFAULT_CASHFLOW_SHEET = "CF 2026 (rolling)"
CASHFLOW_SHEET_FALLBACK = "CF 2026"
CF_MONTH_COLUMNS = {
    22: 1,
    23: 2,
    24: 3,
    25: 4,
    26: 5,
    27: 6,
    28: 7,
    29: 8,
    30: 9,
    31: 10,
    32: 11,
    33: 12,
}
CF_PLAN_EXPENSE_ROWS = range(50, 79)
PROJECT_EXPENSE_PREFIXES = (
    "Projektový náklad - ",
    "Projektové náklady - ",
    "Projektovy naklad - ",
    "Projektove naklady - ",
)
TECHNICAL_EMPLOYEE_NAME = "TENENET Projektové náklady Import"
PROJECT_PLAN_IMPORT_ALLOWLIST = {"ICM"}


def normalize_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def fold_text(value):
    text = unicodedata.normalize("NFKD", normalize_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def slug_text(value):
    return re.sub(r"[^a-z0-9]+", "-", fold_text(value)).strip("-") or "row"


def parse_amount(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        text = normalize_text(value).replace(" ", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0


def month_end(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def split_project_expense_label(label):
    normalized = normalize_text(label)
    for prefix in PROJECT_EXPENSE_PREFIXES:
        if normalized.startswith(prefix):
            return normalized[len(prefix):].strip()
    return ""


def canonicalize_project_row_label(label):
    normalized = normalize_text(label)
    replacements = {
        "Projektovy naklad - ": "Projektový náklad - ",
        "Projektove naklady - ": "Projektové náklady - ",
    }
    for source, target in replacements.items():
        if normalized.startswith(source):
            return target + normalized[len(source):].strip()
    return normalized


def cashflow_plan_row_key(label):
    return f"workbook:expense:{slug_text(label)}"


def parse_project_expense_rows(workbook_path: Path, sheet_name: str = DEFAULT_CASHFLOW_SHEET):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        if sheet_name == DEFAULT_CASHFLOW_SHEET and CASHFLOW_SHEET_FALLBACK in workbook.sheetnames:
            sheet_name = CASHFLOW_SHEET_FALLBACK
        else:
            raise ValueError(f"Cashflow sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    rows = []
    for row_idx in CF_PLAN_EXPENSE_ROWS:
        label = normalize_text(sheet.cell(row_idx, 21).value) or normalize_text(sheet.cell(row_idx, 2).value)
        project_name = split_project_expense_label(label)
        if not project_name:
            continue
        for col_idx, month in CF_MONTH_COLUMNS.items():
            amount = parse_amount(sheet.cell(row_idx, col_idx).value)
            if amount >= 0.0:
                continue
            rows.append({
                "sheet_name": sheet_name,
                "row_idx": row_idx,
                "row_key": cashflow_plan_row_key(label),
                "row_label": label,
                "project_name": project_name,
                "month": month,
                "amount": abs(amount),
            })
    return rows


def _get_or_create_import_employee(env):
    employee = env["hr.employee"].with_context(active_test=False).search(
        [("name", "=", TECHNICAL_EMPLOYEE_NAME)],
        limit=1,
    )
    if employee:
        return employee
    return env["hr.employee"].create({
        "name": TECHNICAL_EMPLOYEE_NAME,
        "work_ratio": 100.0,
    })


def _find_project(env, project_name):
    Project = env["tenenet.project"].with_context(active_test=False)
    direct = Project.search([("display_name", "=", project_name)], limit=1)
    if direct:
        return direct
    direct = Project.search([("name", "=", project_name)], limit=1)
    if direct:
        return direct

    folded_name = fold_text(project_name)
    for candidate in Project.search([]):
        if fold_text(candidate.display_name) == folded_name or fold_text(candidate.name) == folded_name:
            return candidate
    return Project.browse()


def _build_source_key(workbook_path: Path, sheet_name: str, row_idx: int, month: int):
    return f"project_cf_plan:{workbook_path.name}:{sheet_name}:{row_idx}:{month:02d}"


def _is_allowed_project_plan_import(project_name):
    return normalize_text(project_name) in PROJECT_PLAN_IMPORT_ALLOWLIST


def import_project_expenses(env, workbook_path: str | Path, year: int = 2026, sheet_name: str = DEFAULT_CASHFLOW_SHEET):
    workbook_path = Path(workbook_path)
    rows = parse_project_expense_rows(workbook_path, sheet_name=sheet_name)
    employee = _get_or_create_import_employee(env)
    config_model = env["tenenet.expense.type.config"]
    tax = config_model._get_default_operating_tax()
    HrExpense = env["hr.expense"].with_context(mail_create_nosubscribe=True)

    created = 0
    updated = 0
    skipped = 0
    by_project = Counter()
    skipped_labels = Counter()

    for row in rows:
        if not _is_allowed_project_plan_import(row["project_name"]):
            skipped += 1
            skipped_labels[row["row_label"]] += 1
            continue

        project = _find_project(env, row["project_name"])
        if not project:
            skipped += 1
            skipped_labels[row["row_label"]] += 1
            continue

        config = config_model._find_or_create_project_cashflow_import_type(
            row["row_key"],
            canonicalize_project_row_label(row["row_label"]),
        )
        source_key = _build_source_key(workbook_path, row["sheet_name"], row["row_idx"], row["month"])
        notes = (
            f"Import z workbooku {workbook_path.name}, hárok {row['sheet_name']}, riadok {row['row_idx']}, "
            f"mesiac {row['month']:02d}. Pôvodný label: {row['row_label']}"
        )
        values = {
            "name": row["row_label"],
            "description": notes,
            "employee_id": employee.id,
            "date": month_end(year, row["month"]),
            "total_amount_currency": row["amount"],
            "payment_mode": "company_account",
            "tax_ids": [Command.set(tax.ids)],
            "tenenet_cost_flow": "project",
            "tenenet_project_id": project.id,
            "tenenet_expense_type_config_id": config.id,
            "tenenet_add_allowed_type": True,
            "tenenet_allowed_type_limit": 0.0,
            "tenenet_import_source_key": source_key,
        }
        existing = HrExpense.search([("tenenet_import_source_key", "=", source_key)], limit=1)
        if existing:
            existing.write(values)
            updated += 1
        else:
            HrExpense.create(values)
            created += 1
        by_project[project.display_name] += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "rows": len(rows),
        "by_project": dict(by_project),
        "skipped_labels": dict(skipped_labels),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--sheet", default=DEFAULT_CASHFLOW_SHEET)
    args = parser.parse_args()

    if "env" not in globals():
        raise SystemExit("Run this script inside `odoo shell` so `env` is available.")

    result = import_project_expenses(env, args.workbook, year=args.year, sheet_name=args.sheet)
    env.cr.commit()
    print(result)


if __name__ == "__main__":
    main()
