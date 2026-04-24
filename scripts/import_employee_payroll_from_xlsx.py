#!/usr/bin/env python3
from __future__ import annotations

import argparse
import importlib.util
import math
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from odoo import Command, fields


CCP_MULTIPLIER = 1.362
DISABILITY_MULTIPLIERS = {
    "none": CCP_MULTIPLIER,
    "zps": 1.307,
    "tzp": 1.302,
}
PAYROLL_PERIOD_START = date(2026, 1, 1)
ORG_UNIT_BY_EMPLOYER = {
    "tenenet": "TENENET_OZ",
    "scpap": "SCPP",
    "scp&p": "SCPP",
    "kalia": "KALIA",
    "wellnea": "WELLNEA",
}


def _load_ready_import_helpers():
    module_name = "tenenet_import_ready_employees_from_csv"
    if module_name in sys.modules:
        return sys.modules[module_name]
    path = Path(__file__).with_name("import_ready_employees_from_csv.py")
    spec = importlib.util.spec_from_file_location(module_name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


ready_import = _load_ready_import_helpers()
ensure_external_id = ready_import.ensure_external_id
fold_text = ready_import.fold_text
get_or_create_employee_from_ready_row = ready_import.get_or_create_employee
get_or_create_job_from_ready_row = ready_import.get_or_create_job
get_or_create_site = ready_import.get_or_create_site
import_ready_directory = ready_import.import_ready_directory
normalize_text = ready_import.normalize_text
read_csv = ready_import.read_csv
resolve_manager = ready_import.resolve_manager
split_employee_name = ready_import.split_employee_name


@dataclass
class PayrollRow:
    source_row: int
    raw: dict[str, Any]
    name: str
    aliases: set[str]
    gross_12_2025: float | None
    gross_01_2026: float | None
    gross_latest: float | None
    contribution_multiplier: float
    work_hours: float | None
    end_date: date | None
    start_date: date | None
    birthday: date | None

    @property
    def is_active_for_period(self) -> bool:
        return not self.end_date or self.end_date >= PAYROLL_PERIOD_START


@dataclass
class PayrollMerge:
    key: str
    rows: list[PayrollRow] = field(default_factory=list)

    @property
    def active_rows(self) -> list[PayrollRow]:
        return [row for row in self.rows if row.is_active_for_period]

    @property
    def import_rows(self) -> list[PayrollRow]:
        return self.active_rows or self.rows

    @property
    def name(self) -> str:
        return self.primary_row.name

    @property
    def aliases(self) -> set[str]:
        aliases = set()
        for row in self.rows:
            aliases.update(row.aliases)
        return aliases

    @property
    def gross_latest_sum(self) -> float:
        return sum(row.gross_latest or 0.0 for row in self.import_rows if row.gross_latest is not None)

    @property
    def work_hours_sum(self) -> float:
        return sum(row.work_hours or 0.0 for row in self.import_rows if row.work_hours is not None)

    @property
    def primary_row(self) -> PayrollRow:
        rows = self.import_rows
        return sorted(
            rows,
            key=lambda row: (
                row.gross_latest or 0.0,
                row.work_hours or 0.0,
                row.start_date or date.min,
                -row.source_row,
            ),
            reverse=True,
        )[0]

    @property
    def contribution_multiplier(self) -> float:
        primary = self.primary_row
        if primary.contribution_multiplier:
            return primary.contribution_multiplier
        return CCP_MULTIPLIER

    @property
    def changed_salary_rows(self) -> list[PayrollRow]:
        return [
            row for row in self.rows
            if row.gross_12_2025 is not None
            and row.gross_01_2026 is not None
            and abs(row.gross_12_2025 - row.gross_01_2026) > 0.005
        ]


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = normalize_text(str(value)).replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(str(value))
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def fold_alias(value: str) -> str:
    text = normalize_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def strip_parenthetical(value: str) -> str:
    result = []
    depth = 0
    for char in value:
        if char == "(":
            depth += 1
            continue
        if char == ")":
            depth = max(0, depth - 1)
            continue
        if not depth:
            result.append(char)
    return normalize_text("".join(result))


def name_aliases(name: str) -> set[str]:
    normalized = normalize_text(name)
    aliases = {fold_alias(normalized)} if normalized else set()
    without_parentheses = strip_parenthetical(normalized)
    if without_parentheses:
        aliases.add(fold_alias(without_parentheses))

    for candidate in list(aliases):
        parts = candidate.split()
        if len(parts) == 2:
            aliases.add(f"{parts[1]} {parts[0]}")

    display_parts = without_parentheses.split()
    if len(display_parts) >= 2 and "/" in display_parts[0]:
        rest = " ".join(display_parts[1:])
        for surname in display_parts[0].split("/"):
            aliases.update(name_aliases(f"{surname} {rest}"))
    return {alias for alias in aliases if alias}


def canonical_key(name: str) -> str:
    aliases = sorted(name_aliases(name), key=lambda value: (len(value), value))
    return aliases[0] if aliases else fold_alias(name)


def org_code_from_employer(value: Any) -> str:
    key = fold_alias(str(value or ""))
    return ORG_UNIT_BY_EMPLOYER.get(key, "TENENET_OZ")


def get_organizational_unit(env, employer: Any):
    return env["tenenet.organizational.unit"].search([("code", "=", org_code_from_employer(employer))], limit=1)


def education_value(row: PayrollRow) -> str:
    school = normalize_text(row.raw.get("Vzdelanie") or "")
    field = normalize_text(row.raw.get("Odbor") or "")
    if school and field:
        return f"{school} - {field}"
    return school or field


def choose_non_empty(rows: list[PayrollRow], getter):
    for row in sorted(rows, key=lambda item: (item.start_date or date.min, item.source_row), reverse=True):
        value = getter(row)
        if normalize_text(str(value or "")):
            return value
    return None


def disability_type_from_source(disability_label: Any, multiplier: float | None) -> str:
    if multiplier is not None:
        if math.isclose(multiplier, DISABILITY_MULTIPLIERS["tzp"], rel_tol=0.0, abs_tol=0.0005):
            return "tzp"
        if math.isclose(multiplier, DISABILITY_MULTIPLIERS["zps"], rel_tol=0.0, abs_tol=0.0005):
            return "zps"
    label = fold_text(normalize_text(disability_label or "")).lower()
    if "tzp" in label:
        return "tzp"
    if "zps" in label:
        return "zps"
    return "none"


def read_payroll_rows(xlsx_path: Path, sheet_name: str = "all") -> list[PayrollRow]:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[PayrollRow] = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(headers, values))
        name = normalize_text(raw.get("Priezvisko a meno") or "")
        if not name:
            continue
        gross_01 = parse_float(raw.get("Mzda 01/2026"))
        gross_12 = parse_float(raw.get("Mzda 12/2025"))
        contribution = parse_float(raw.get("Odvody")) or CCP_MULTIPLIER
        rows.append(PayrollRow(
            source_row=index,
            raw=raw,
            name=name,
            aliases=name_aliases(name),
            gross_12_2025=gross_12,
            gross_01_2026=gross_01,
            gross_latest=gross_01 if gross_01 is not None else gross_12,
            contribution_multiplier=contribution,
            work_hours=parse_float(raw.get("Úväzok (hod)")),
            end_date=parse_date(raw.get("Dátum ukončenia PP")),
            start_date=parse_date(raw.get("Dátum nástupu do zamestnania")),
            birthday=parse_date(raw.get("Dátum nar.")),
        ))
    return rows


def merge_payroll_rows(rows: list[PayrollRow]) -> list[PayrollMerge]:
    groups: dict[str, PayrollMerge] = {}
    for row in rows:
        key = canonical_key(row.name)
        groups.setdefault(key, PayrollMerge(key=key)).rows.append(row)
    return list(groups.values())


def load_ready_employee_maps(csv_dir: Path) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    ready_rows = read_csv(csv_dir / "hr_employee_import_ready.csv")
    by_alias = {}
    employee_xml_by_alias = {}
    for row in ready_rows:
        aliases = name_aliases(row.get("name") or "")
        if row.get("work_email"):
            aliases.add(fold_text(row["work_email"]))
        for alias in aliases:
            by_alias[alias] = row
            employee_xml_by_alias[alias] = row["id"]
    return by_alias, employee_xml_by_alias


def find_ready_row(merge: PayrollMerge, ready_by_alias: dict[str, dict[str, str]]) -> dict[str, str] | None:
    email = normalize_text(str(choose_non_empty(merge.rows, lambda row: row.raw.get("Email")) or ""))
    if email and fold_text(email) in ready_by_alias:
        return ready_by_alias[fold_text(email)]
    for alias in merge.aliases:
        if alias in ready_by_alias:
            return ready_by_alias[alias]
    return None


def employee_search_aliases(employee) -> set[str]:
    aliases = set()
    for value in [
        employee.name,
        employee.legal_name,
        " ".join(part for part in [employee.first_name or "", employee.last_name or ""] if part),
        " ".join(part for part in [employee.last_name or "", employee.first_name or ""] if part),
    ]:
        aliases.update(name_aliases(value or ""))
    return aliases


def find_employee(env, merge: PayrollMerge, ready_row: dict[str, str] | None, employee_map: dict[str, int]):
    Employee = env["hr.employee"].with_context(active_test=False)
    if ready_row and ready_row.get("id") in employee_map:
        employee = Employee.browse(employee_map[ready_row["id"]]).exists()
        if employee:
            return employee
    email = normalize_text(str(choose_non_empty(merge.rows, lambda row: row.raw.get("Email")) or ""))
    if email:
        employee = Employee.search([("work_email", "=", email)], limit=1)
        if employee:
            return employee
    for employee in Employee.search([]):
        if merge.aliases.intersection(employee_search_aliases(employee)):
            return employee
    return Employee


def get_or_create_job(env, name: str):
    name = normalize_text(name)
    if not name:
        return env["hr.job"]
    job = env["hr.job"].search([("name", "=", name)], limit=1)
    if job:
        return job
    xml_id = "job_" + "".join(
        char if char.isalnum() else "_"
        for char in fold_text(name).replace(" ", "_")
    ).strip("_")
    return get_or_create_job_from_ready_row(env, {"id": xml_id, "name": name})


def build_employee_values(env, merge: PayrollMerge, ready_row: dict[str, str] | None) -> dict[str, Any]:
    primary = merge.primary_row
    first_name, last_name = split_employee_name(ready_row["name"] if ready_row else primary.name)
    org_unit = get_organizational_unit(env, primary.raw.get("Zamestnávateľ"))
    latest_gross = merge.gross_latest_sum
    multiplier = merge.contribution_multiplier
    education = choose_non_empty(merge.rows, education_value)
    birthday = choose_non_empty(merge.rows, lambda row: row.birthday)
    start_date = choose_non_empty(merge.rows, lambda row: row.start_date)
    end_dates = [row.end_date for row in merge.import_rows if row.end_date]
    end_date = False if any(not row.end_date for row in merge.import_rows) else (max(end_dates) if end_dates else False)
    disabled_state = choose_non_empty(merge.rows, lambda row: row.raw.get("ZPS/ŤZP"))
    disability_type = disability_type_from_source(disabled_state, multiplier)
    is_disabled = disability_type != "none"
    email = choose_non_empty(merge.rows, lambda row: row.raw.get("Email"))
    position = normalize_text(primary.raw.get("Pozícia") or primary.raw.get("Pozícia podľa pracovnej zmluvy") or "")
    job = get_or_create_job(env, position)
    site = get_or_create_site(env, normalize_text(primary.raw.get("Lokácia") or primary.raw.get("Miesto výkonu") or ""))
    vals = {
        "name": ready_row["name"] if ready_row else primary.name,
        "first_name": first_name or False,
        "last_name": last_name or False,
        "title_academic": normalize_text(primary.raw.get("Titul") or "") or (ready_row.get("title_academic") if ready_row else False),
        "organizational_unit_id": org_unit.id or False,
        "birthday": birthday or False,
        "contract_date_start": start_date or False,
        "contract_date_end": end_date or False,
        "experience_years_total": parse_float(primary.raw.get("Počet rokov praxe k 01.01.2025")) or 0.0,
        "work_ratio": (merge.work_hours_sum / 8.0 * 100.0) if merge.work_hours_sum else 0.0,
        "tenenet_payroll_contribution_multiplier": multiplier,
        "tenenet_disability_type": disability_type,
        "monthly_gross_salary_target": latest_gross * multiplier if latest_gross else 0.0,
        "education_info": education or False,
        "contract_position": normalize_text(primary.raw.get("Pozícia podľa pracovnej zmluvy") or "") or False,
        "position": position or False,
        "work_email": normalize_text(email or "") or False,
        "job_id": job.id or False,
        "position_catalog_id": job.id or False,
        "main_site_id": site.id if site else False,
        "secondary_site_ids": [Command.set([])],
        "work_location_id": False,
        "address_id": False,
    }
    if "disabled" in env["hr.employee"]._fields:
        vals["disabled"] = is_disabled
    return vals


def has_importable_payroll_data(merge: PayrollMerge) -> bool:
    if merge.gross_latest_sum or merge.work_hours_sum:
        return True
    return any(
        row.birthday or row.start_date or education_value(row) or normalize_text(row.raw.get("Email") or "")
        for row in merge.rows
    )


def update_employee_from_payroll(env, merge: PayrollMerge, ready_row: dict[str, str] | None, employee_map: dict[str, int]):
    employee = find_employee(env, merge, ready_row, employee_map)
    vals = build_employee_values(env, merge, ready_row)
    created = False
    if employee:
        employee.write(vals)
    else:
        employee = env["hr.employee"].create(vals)
        created = True

    xml_name = ready_row["id"] if ready_row else "emp_" + "_".join(fold_text(vals["name"]).split())
    ensure_external_id(env, employee, xml_name)
    if ready_row:
        employee_map[ready_row["id"]] = employee.id
    return employee, created


def write_report(report_path: Path, report: dict[str, Any]) -> None:
    lines = [
        "# Employee Payroll Excel Import Review",
        "",
        f"- Source Excel rows: {report['source_rows']}",
        f"- Unique merged employees: {report['merged_employees']}",
        f"- Created employees: {len(report['created'])}",
        f"- Updated employees: {len(report['updated'])}",
        f"- Skipped groups: {len(report['skipped'])}",
        f"- Duplicate merge groups: {len(report['duplicate_merges'])}",
        f"- Salary changes detected: {len(report['salary_changes'])}",
        f"- Missing education after merge: {len(report['missing_education'])}",
        f"- Ambiguous org units: {len(report['ambiguous_org_units'])}",
        "",
    ]
    for title, rows in [
        ("Created Employees", report["created"]),
        ("Updated Employees", report["updated"]),
        ("Duplicate Merges", report["duplicate_merges"]),
        ("Salary Changes", report["salary_changes"]),
        ("Ambiguous Org Units", report["ambiguous_org_units"]),
        ("Missing Education", report["missing_education"]),
        ("Skipped", report["skipped"]),
    ]:
        lines.extend([f"## {title}", ""])
        if rows:
            lines.extend(f"- {row}" for row in rows)
        else:
            lines.append("- none")
        lines.append("")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines), encoding="utf-8")


def import_payroll_xlsx(env, xlsx_path: Path, ready_dir: Path, report_path: Path | None = None) -> dict[str, Any]:
    base_result = import_ready_directory(env, ready_dir)
    ready_by_alias, _employee_xml_by_alias = load_ready_employee_maps(ready_dir)
    ready_rows = read_csv(ready_dir / "hr_employee_import_ready.csv")
    employee_map = {
        row["id"]: env["ir.model.data"]._xmlid_to_res_id(f"tenenet_projects_import.{row['id']}", raise_if_not_found=False)
        for row in ready_rows
    }
    employee_map = {key: value for key, value in employee_map.items() if value}

    payroll_rows = read_payroll_rows(xlsx_path)
    merges = merge_payroll_rows(payroll_rows)
    report = {
        "base_result": base_result,
        "source_rows": len(payroll_rows),
        "merged_employees": len(merges),
        "created": [],
        "updated": [],
        "skipped": [],
        "duplicate_merges": [],
        "salary_changes": [],
        "ambiguous_org_units": [],
        "missing_education": [],
    }

    for merge in merges:
        if not has_importable_payroll_data(merge):
            report["skipped"].append(f"{merge.name} | rows {[row.source_row for row in merge.rows]} | no usable payroll/profile data")
            continue

        ready_row = find_ready_row(merge, ready_by_alias)
        if len(merge.rows) > 1:
            report["duplicate_merges"].append(
                f"{merge.name} | rows {[row.source_row for row in merge.rows]} | hours {merge.work_hours_sum:g} | gross {merge.gross_latest_sum:g}"
            )
        org_codes = {org_code_from_employer(row.raw.get("Zamestnávateľ")) for row in merge.import_rows}
        if len(org_codes) > 1:
            report["ambiguous_org_units"].append(
                f"{merge.name} | rows {[row.source_row for row in merge.rows]} | units {', '.join(sorted(org_codes))} | chosen {org_code_from_employer(merge.primary_row.raw.get('Zamestnávateľ'))}"
            )
        if not choose_non_empty(merge.rows, education_value):
            report["missing_education"].append(f"{merge.name} | rows {[row.source_row for row in merge.rows]}")
        for row in merge.changed_salary_rows:
            report["salary_changes"].append(
                f"{row.name} | row {row.source_row} | {row.gross_12_2025:g} -> {row.gross_01_2026:g}"
            )

        employee, created = update_employee_from_payroll(env, merge, ready_row, employee_map)
        label = f"{employee.display_name} | rows {[row.source_row for row in merge.rows]} | CCP {employee.monthly_gross_salary_target:g}"
        report["created" if created else "updated"].append(label)

    if report_path:
        write_report(report_path, report)
    return report


def main() -> None:
    current_env = globals().get("env")
    if current_env is None:
        raise RuntimeError(
            "Odoo env is not available. Run this script from `odoo-bin shell` and pass `init_globals={'env': env}` to runpy.run_path()."
        )

    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("ready_dir", type=Path)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    report = import_payroll_xlsx(current_env, args.xlsx_path, args.ready_dir, args.report)
    current_env.cr.commit()
    print(f"BASE_EMPLOYEES={report['base_result']['employees']}")
    print(f"PAYROLL_ROWS={report['source_rows']}")
    print(f"MERGED_EMPLOYEES={report['merged_employees']}")
    print(f"CREATED={len(report['created'])}")
    print(f"UPDATED={len(report['updated'])}")
    print(f"SKIPPED={len(report['skipped'])}")


if __name__ == "__main__":
    main()
