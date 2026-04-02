#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


SOURCE_HEADERS = [
    "Titul",
    "Priezvisko a meno",
    "Program,do ktorého patria",
    "Pozícia podľa pracovnej zmluvy",
    "Pozícia",
    "Priamy nadriadený",
    "Lokácia",
    "Email",
    "TEL",
]

EMPLOYEE_HEADERS = [
    "id",
    "title_academic",
    "name",
    "job_id/id",
    "department_id/id",
    "parent_id/id",
    "address_id/id",
    "work_location",
    "work_email",
    "mobile_phone",
    "work_phone",
    "x_source_row",
    "x_raw_program",
    "x_program_normalized",
    "x_raw_contract_position",
    "x_raw_position",
    "x_normalization_status",
    "x_review_note",
]

JOB_HEADERS = ["id", "name", "x_department_code"]
DEPARTMENT_HEADERS = ["id", "name"]
LOCATION_HEADERS = ["id", "name", "type", "city", "country_id/id"]
PROGRAM_HEADERS = ["raw_program", "normalized_program_code", "normalized_program_name", "classification"]

INVALID_PHONE_MARKERS = {
    "",
    "x",
    "n/a",
    "na",
}

PROGRAM_LOOKUP = {
    "AKP": ("AKP_SHARED", "AKP"),
    "AKP - DETI A MLADEZ": ("AKP_DETI", "AKP - deti a mládež"),
    "APZ": ("APZ", "APZ"),
    "APZ, EU CARE": ("APZ_EU_CARE", "APZ, EU CARE"),
    "AVL": ("AVL", "AVL"),
    "EU CARE": ("EU_CARE", "EU CARE"),
    "KC - KOMUNITNE CENTRA": ("KC", "Komunitné centrum"),
    "NAS A VAZENSTVO": ("NAS_A_VAZ", "Násilie a väzenstvo"),
    "PSC BB": ("PSC_BB", "Psychiatrické centrum Banská Bystrica"),
    "PSC KOSICE": ("PSC_KE", "Psychiatrické centrum Košice"),
    "PSC SENEC": ("PSC_SC", "Psychiatrické centrum Senec"),
    "SCPAP": ("SCPAP", "Súkromné centrum poradenstva a prevencie"),
    "SCPP": ("SCPP", "ŠCPP"),
    "SPODASK": ("SPODASK", "SPODaSK"),
    "STEM, GUIDE, SOCIALNE INOVACIE, APZ": ("SUPPORT_PROJECTS", "STEM, GUIDE, Sociálne inovácie, APZ"),
    "SVI": ("VCI", "Včasná intervencia"),
    "SSP": ("SSP", "Špecializované sociálne poradenstvo"),
}

PROGRAM_NOISE_RULES = [
    (re.compile(r"\bned[aá]va[ťt]\b", re.IGNORECASE), ("NO_IMPORT", "Poznámka - neimportovať")),
    (re.compile(r"u[žz]\s+je\s+na\s+webe", re.IGNORECASE), ("WEB_ONLY", "Poznámka - už je na webe")),
    (re.compile(r"kon[cč][ií].*2026", re.IGNORECASE), ("TRANSITION_NOTE", "Poznámka - prechod programu")),
    (re.compile(r"lep[sš]ie\s+pomenovanie.*SVI", re.IGNORECASE), ("VCI", "Včasná intervencia")),
]

LOCATION_FIXES = {
    "Bratsislava": "Bratislava",
    "Galanta Trnava": "Galanta, Trnava",
    "Senec BSK, TTSK": "Senec, BSK, TTSK",
}

JOB_SPECS = {
    "psycholog": ("Psychológ", "psychologia"),
    "specialny_pedagog": ("Špeciálny pedagóg", "specialna_pedagogika_terapie"),
    "liecebny_pedagog": ("Liečebný pedagóg", "specialna_pedagogika_terapie"),
    "socialny_pracovnik": ("Sociálny pracovník", "socialne_sluzby"),
    "komunitny_pracovnik": ("Komunitný pracovník", "komunitne_centrum"),
    "krizova_intervencia": ("Pracovník krízovej intervencie", "socialne_sluzby"),
    "zdravotna_sestra": ("Zdravotná sestra", "zdravotna_starostlivost"),
    "zdravotna_sestra_psychiatria": ("Zdravotná sestra v psychiatrii", "zdravotna_starostlivost"),
    "zdravotnicky_asistent": ("Zdravotnícky asistent", "zdravotna_starostlivost"),
    "lekar": ("Lekár", "zdravotna_starostlivost"),
    "psychiater": ("Psychiater", "zdravotna_starostlivost"),
    "logoped": ("Logopéd", "specialna_pedagogika_terapie"),
    "fyzioterapeut": ("Fyzioterapeut", "specialna_pedagogika_terapie"),
    "odborny_garant": ("Odborný garant", "manazment"),
    "odborny_riaditel": ("Odborný riaditeľ", "manazment"),
    "programovy_riaditel": ("Programový riaditeľ", "manazment"),
    "generalna_riaditelka": ("Generálna riaditeľka", "manazment"),
    "financna_riaditelka": ("Finančná riaditeľka", "manazment"),
    "financny_manazer": ("Finančný manažér", "prevadzka_a_podpora"),
    "projektovy_manazer": ("Projektový manažér", "prevadzka_a_podpora"),
    "prevadzkova_manazerka": ("Prevádzková manažérka", "prevadzka_a_podpora"),
    "personalista": ("Personalista", "prevadzka_a_podpora"),
    "mzdova_uctovnicka": ("Mzdová účtovníčka", "prevadzka_a_podpora"),
    "uctovnicka": ("Účtovníčka", "prevadzka_a_podpora"),
    "recepcna": ("Recepčná", "prevadzka_a_podpora"),
    "upratovacka": ("Upratovačka", "prevadzka_a_podpora"),
    "vyskumno_vyvojovy_pracovnik": ("Výskumný/vývojový pracovník", "prevadzka_a_podpora"),
    "socialna_inkluzia_pracovnik": ("Odborný pracovník pre sociálnu inklúziu a zamestnanosť", "socialne_sluzby"),
}

DEPARTMENTS = {
    "psychologia": "Psychológia",
    "socialne_sluzby": "Sociálne služby",
    "zdravotna_starostlivost": "Zdravotná starostlivosť",
    "specialna_pedagogika_terapie": "Špeciálna pedagogika a terapie",
    "komunitne_centrum": "Komunitné centrum",
    "prevadzka_a_podpora": "Prevádzka a podpora",
    "manazment": "Manažment",
}


@dataclass
class CleanedRow:
    source_row: int
    title: str
    name: str
    raw_program: str
    raw_contract_position: str
    raw_position: str
    raw_manager: str
    raw_location: str
    raw_email: str
    raw_phone: str
    program_code: str
    program_name: str
    program_classification: str
    job_key: str
    job_name: str
    department_key: str
    department_name: str
    manager_name: str
    manager_employee_id: str
    work_location: str
    location_partner_id: str
    work_email: str
    mobile_phone: str
    work_phone: str
    status: str
    review_note: str
    employee_id: str


def normalize_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def fold_text(value: str) -> str:
    text = normalize_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def slugify(value: str) -> str:
    text = fold_text(value)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def employee_xmlid(name: str) -> str:
    return f"emp_{slugify(name)}"


def job_xmlid(job_key: str) -> str:
    return f"job_{job_key}"


def department_xmlid(department_key: str) -> str:
    return f"dept_{department_key}"


def location_xmlid(city: str) -> str:
    return f"loc_{slugify(city)}"


def sentence_case(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return text[:1].upper() + text[1:].lower()


def classify_program(raw_program: str) -> tuple[str, str, str]:
    value = normalize_text(raw_program)
    if not value:
        return "", "", "empty"

    for pattern, result in PROGRAM_NOISE_RULES:
        if pattern.search(value):
            code, label = result
            if code == "VCI":
                return code, label, "normalized_from_note"
            return code, label, "note"

    lookup_key = fold_text(value).upper()
    if lookup_key in PROGRAM_LOOKUP:
        code, label = PROGRAM_LOOKUP[lookup_key]
        return code, label, "mapped"
    return "", "", "unmapped"


def normalize_location(raw_location: str) -> str:
    value = normalize_text(raw_location)
    if not value:
        return ""
    return LOCATION_FIXES.get(value, value)


def split_locations(location: str) -> list[str]:
    normalized = normalize_location(location)
    if not normalized:
        return []
    return [part.strip() for part in normalized.split(",") if part.strip()]


def clean_email(raw_email: str) -> tuple[str, list[str]]:
    notes: list[str] = []
    email = normalize_text(raw_email).lower()
    if not email:
        return "", notes
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]{2,}", email):
        notes.append("invalid email blanked")
        return "", notes
    return email, notes


def format_phone(digits: str) -> str:
    if len(digits) == 9 and digits.startswith("9"):
        return f"+421 {digits[0:3]} {digits[3:6]} {digits[6:9]}"
    return digits


def format_landline(national_number: str) -> str:
    if len(national_number) == 8 and national_number.startswith("2"):
        return f"+421 2 {national_number[1:5]} {national_number[5:9]}"
    if len(national_number) == 9:
        return f"+421 {national_number[0:2]} {national_number[2:5]} {national_number[5:7]} {national_number[7:9]}"
    if len(national_number) == 8:
        return f"+421 {national_number[0:2]} {national_number[2:5]} {national_number[5:8]}"
    return f"+421 {national_number}"


def clean_phone(raw_phone: str) -> tuple[str, str, list[str]]:
    notes: list[str] = []
    phone = normalize_text(raw_phone)
    if not phone:
        return "", "", notes

    folded = fold_text(phone)
    if folded in INVALID_PHONE_MARKERS:
        notes.append("invalid phone blanked")
        return "", "", notes
    if "recepcie" in folded or "recepcia" in folded:
        notes.append("shared reception phone omitted")
        return "", "", notes

    digits = re.sub(r"\D", "", phone)
    if not digits:
        notes.append("invalid phone blanked")
        return "", "", notes

    if len(digits) == 9 and digits.startswith("9"):
        return format_phone(digits), "", notes
    if len(digits) == 9 and digits[0] in "2345678":
        return "", format_landline(digits), notes
    if len(digits) == 10 and digits.startswith("0"):
        return "", format_landline(digits[1:]), notes

    notes.append("unrecognized phone format blanked")
    return "", "", notes


def resolve_job(raw_position: str, raw_contract_position: str, program_code: str) -> tuple[str, list[str]]:
    position = fold_text(raw_position)
    contract = fold_text(raw_contract_position)
    notes: list[str] = []

    if position == "psycholog" or contract == "psycholog":
        return "psycholog", notes
    if position == "specialny pedagog":
        return "specialny_pedagog", notes
    if position == "liecebny pedagog" or contract == "liecebny pedagog":
        notes.append("position normalized from typo liečebný pedágóg")
        return "liecebny_pedagog", notes
    if position == "socialny pracovnik":
        return "socialny_pracovnik", notes
    if position in {"komunitny pracovnik", "pracovnik kc"} or contract in {"komunitny pracovnik", "pracovnik kc"}:
        if position != "komunitny pracovnik":
            notes.append("position normalized to komunitný pracovník")
        return "komunitny_pracovnik", notes
    if position == "pracovnik krizovej intervencie":
        return "krizova_intervencia", notes
    if position == "zdravotna sestra":
        return "zdravotna_sestra", notes
    if position == "zdravotna sestra v psychiatrii":
        return "zdravotna_sestra_psychiatria", notes
    if position == "zdravotnicky asistent":
        return "zdravotnicky_asistent", notes
    if position in {"lekar"} or contract == "vseobecny lekar pre dospelych":
        return "lekar", notes
    if position == "psychiater" or contract == "psychiater":
        return "psychiater", notes
    if position == "zdravotna sestra/zdravotny brat" or contract == "zdravotna sestra/zdravotny brat":
        notes.append("position normalized to zdravotná sestra")
        return "zdravotna_sestra", notes
    if position == "logoped":
        return "logoped", notes
    if position == "fyzioterapeut":
        return "fyzioterapeut", notes
    if position == "odborny riaditel" or "odborny riaditel" in contract:
        return "odborny_riaditel", notes
    if position == "odborny garant":
        return "odborny_garant", notes
    if position.startswith("odborny garant pre ambulanciu psychiatra"):
        notes.append("position collapsed to odborný garant")
        return "odborny_garant", notes
    if position.startswith("odborny garant pre klinicku psychologiu"):
        notes.append("position collapsed to odborný garant")
        return "odborny_garant", notes
    if position == "odborny garant spodask" or "odborny garant spodask" in contract:
        notes.append("position collapsed to odborný garant")
        return "odborny_garant", notes
    if position == "programovy riaditel" or "programovy riaditel" in contract:
        return "programovy_riaditel", notes
    if position == "generalna riaditelka" or "generalna riaditelka" in contract:
        return "generalna_riaditelka", notes
    if position == "financna riaditelka" or contract == "financny riaditel":
        return "financna_riaditelka", notes
    if position == "financny manazer" or contract == "financny manazer":
        return "financny_manazer", notes
    if position == "projektovy manazer" or contract == "projektovy manazer":
        return "projektovy_manazer", notes
    if position == "apz":
        notes.append("position inferred to projektový manažér from APZ placeholder")
        return "projektovy_manazer", notes
    if position == "prevadzkova manazerka / odborna garantka podprogramu pre osoby so zdravotnym znevyhodnenim":
        notes.append("position shortened to prevádzková manažérka")
        return "prevadzkova_manazerka", notes
    if position == "personalista":
        return "personalista", notes
    if position == "mzdova uctovnicka":
        return "mzdova_uctovnicka", notes
    if position == "uctovnicka":
        return "uctovnicka", notes
    if position == "recepcna":
        return "recepcna", notes
    if position == "upratovacka":
        return "upratovacka", notes
    if position == "vyskumny/vyvojovy pracovnik":
        return "vyskumno_vyvojovy_pracovnik", notes
    if contract.startswith("odborny/a pracovnik/cka pre socialnu inkluziu"):
        notes.append("position inferred from contract text")
        return "socialna_inkluzia_pracovnik", notes

    raise ValueError(f"Unmapped position: position={raw_position!r}, contract={raw_contract_position!r}, program={program_code!r}")


def canonical_manager_aliases(name: str) -> set[str]:
    normalized = normalize_text(name)
    folded = fold_text(normalized)
    aliases = {folded}
    parts = normalized.split()
    if len(parts) == 2:
        aliases.add(f"{fold_text(parts[1])} {fold_text(parts[0])}")
    return aliases


def build_manager_lookup(employee_names: list[str]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for name in employee_names:
        employee_id = employee_xmlid(name)
        for alias in canonical_manager_aliases(name):
            lookup[alias] = employee_id
    return lookup


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def collect_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["all"]
    rows: list[dict[str, str]] = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        selected = row[1:10]
        if not any(selected):
            continue
        record = dict(zip(SOURCE_HEADERS, selected))
        record["__row__"] = row_idx
        rows.append(record)
    return rows


def clean_rows(source_rows: list[dict[str, str]]) -> list[CleanedRow]:
    employee_names = [normalize_text(row["Priezvisko a meno"]) for row in source_rows]
    manager_lookup = build_manager_lookup(employee_names)
    cleaned: list[CleanedRow] = []

    for row in source_rows:
        source_row = row["__row__"]
        title = normalize_text(row["Titul"])
        name = normalize_text(row["Priezvisko a meno"])
        raw_program = normalize_text(row["Program,do ktorého patria"])
        raw_contract_position = normalize_text(row["Pozícia podľa pracovnej zmluvy"])
        raw_position = normalize_text(row["Pozícia"])
        raw_manager = normalize_text(row["Priamy nadriadený"])
        raw_location = normalize_text(row["Lokácia"])
        raw_email = normalize_text(row["Email"])
        raw_phone = normalize_text(row["TEL"])

        review_notes: list[str] = []
        program_code, program_name, program_classification = classify_program(raw_program)
        if program_classification in {"note", "normalized_from_note", "unmapped"}:
            review_notes.append(f"program flagged: {raw_program or 'empty'}")

        job_key, job_notes = resolve_job(raw_position, raw_contract_position, program_code)
        review_notes.extend(job_notes)
        job_name, department_key = JOB_SPECS[job_key]
        job_name = sentence_case(job_name)
        department_name = sentence_case(DEPARTMENTS[department_key])

        location = normalize_location(raw_location)
        if location != raw_location and raw_location:
            review_notes.append(f"location normalized from {raw_location}")
        location_parts = split_locations(location)
        location_partner_id = ""
        if len(location_parts) == 1:
            location_partner_id = location_xmlid(location_parts[0])
        elif len(location_parts) > 1:
            review_notes.append(f"multiple locations: {', '.join(location_parts)}")

        email, email_notes = clean_email(raw_email)
        review_notes.extend(email_notes)

        mobile_phone, work_phone, phone_notes = clean_phone(raw_phone)
        review_notes.extend(phone_notes)

        manager_name = ""
        manager_employee_id = ""
        if raw_manager and fold_text(raw_manager) not in {"n/a", "na"}:
            manager_id = manager_lookup.get(fold_text(raw_manager))
            if not manager_id:
                parts = normalize_text(raw_manager).split()
                if len(parts) == 2:
                    manager_id = manager_lookup.get(f"{fold_text(parts[1])} {fold_text(parts[0])}")
                    if manager_id:
                        review_notes.append(f"manager normalized from {raw_manager}")
            if manager_id:
                manager_employee_id = manager_id
                manager_name = next(name for name in employee_names if employee_xmlid(name) == manager_id)
            else:
                review_notes.append(f"unresolved manager: {raw_manager}")
        elif raw_manager:
            review_notes.append("manager omitted from N/A marker")

        status = "ready"
        if review_notes:
            status = "review"

        cleaned.append(
            CleanedRow(
                source_row=source_row,
                title=title,
                name=name,
                raw_program=raw_program,
                raw_contract_position=raw_contract_position,
                raw_position=raw_position,
                raw_manager=raw_manager,
                raw_location=raw_location,
                raw_email=raw_email,
                raw_phone=raw_phone,
                program_code=program_code,
                program_name=program_name,
                program_classification=program_classification,
                job_key=job_key,
                job_name=job_name,
                department_key=department_key,
                department_name=department_name,
                manager_name=manager_name,
                manager_employee_id=manager_employee_id,
                work_location=location,
                location_partner_id=location_partner_id,
                work_email=email,
                mobile_phone=mobile_phone,
                work_phone=work_phone,
                status=status,
                review_note="; ".join(review_notes),
                employee_id=employee_xmlid(name),
            )
        )

    return cleaned


def export_outputs(cleaned_rows: list[CleanedRow], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    employee_rows = [
        {
            "id": row.employee_id,
            "title_academic": row.title,
            "name": row.name,
            "job_id/id": job_xmlid(row.job_key),
            "department_id/id": department_xmlid(row.department_key),
            "parent_id/id": row.manager_employee_id,
            "address_id/id": row.location_partner_id,
            "work_location": row.work_location,
            "work_email": row.work_email,
            "mobile_phone": row.mobile_phone,
            "work_phone": row.work_phone,
            "x_source_row": str(row.source_row),
            "x_raw_program": row.raw_program,
            "x_program_normalized": row.program_code or row.program_name,
            "x_raw_contract_position": row.raw_contract_position,
            "x_raw_position": row.raw_position,
            "x_normalization_status": row.status,
            "x_review_note": row.review_note,
        }
        for row in cleaned_rows
    ]

    job_rows = [
        {
            "id": job_xmlid(job_key),
            "name": job_name,
            "x_department_code": department_xmlid(department_key),
        }
        for job_key, (job_name, department_key) in sorted(JOB_SPECS.items(), key=lambda item: item[1][0])
        if any(row.job_key == job_key for row in cleaned_rows)
    ]

    department_rows = [
        {"id": department_xmlid(department_key), "name": department_name}
        for department_key, department_name in sorted(DEPARTMENTS.items(), key=lambda item: item[1])
        if any(row.department_key == department_key for row in cleaned_rows)
    ]

    location_values = sorted({
        row.work_location
        for row in cleaned_rows
        if row.location_partner_id and row.work_location
    })
    location_rows = [
        {
            "id": location_xmlid(city),
            "name": f"Work location - {city}",
            "type": "other",
            "city": city,
            "country_id/id": "base.sk",
        }
        for city in location_values
    ]

    program_rows = []
    seen_programs = set()
    for row in sorted(cleaned_rows, key=lambda item: (item.raw_program, item.source_row)):
        if row.raw_program in seen_programs:
            continue
        seen_programs.add(row.raw_program)
        program_rows.append(
            {
                "raw_program": row.raw_program,
                "normalized_program_code": row.program_code,
                "normalized_program_name": row.program_name,
                "classification": row.program_classification,
            }
        )

    employee_path = output_dir / "hr_employee_import.csv"
    job_path = output_dir / "hr_job_import.csv"
    department_path = output_dir / "hr_department_import.csv"
    location_path = output_dir / "res_partner_location_import.csv"
    program_path = output_dir / "program_normalization.csv"
    review_path = output_dir / "employee_cleanup_review.md"

    write_csv(employee_path, EMPLOYEE_HEADERS, employee_rows)
    write_csv(job_path, JOB_HEADERS, job_rows)
    write_csv(department_path, DEPARTMENT_HEADERS, department_rows)
    write_csv(location_path, LOCATION_HEADERS, location_rows)
    write_csv(program_path, PROGRAM_HEADERS, program_rows)
    review_path.write_text(build_review_report(cleaned_rows), encoding="utf-8")

    return {
        "employee": employee_path,
        "job": job_path,
        "department": department_path,
        "location": location_path,
        "program": program_path,
        "review": review_path,
    }


def build_review_report(cleaned_rows: list[CleanedRow]) -> str:
    review_rows = [row for row in cleaned_rows if row.review_note]
    unresolved_managers = [row for row in cleaned_rows if "unresolved manager:" in row.review_note or "manager omitted" in row.review_note]
    invalid_contacts = [row for row in cleaned_rows if "email blanked" in row.review_note or "phone" in row.review_note]
    noisy_programs = [row for row in cleaned_rows if "program flagged:" in row.review_note]
    position_judgments = [row for row in cleaned_rows if "position " in row.review_note]
    multi_locations = [row for row in cleaned_rows if "multiple locations:" in row.review_note]

    sections = [
        "# Employee Cleanup Review",
        "",
        f"- Source rows processed: {len(cleaned_rows)}",
        f"- Rows requiring review: {len(review_rows)}",
        f"- Unresolved or omitted managers: {len(unresolved_managers)}",
        f"- Contact cleanup issues: {len(invalid_contacts)}",
        f"- Program notes/noise rows: {len(noisy_programs)}",
        f"- Non-trivial position normalizations: {len(position_judgments)}",
        f"- Rows with multiple locations: {len(multi_locations)}",
        "",
        "## Rows Requiring Review",
        "",
    ]

    for row in review_rows:
        sections.append(
            f"- Row {row.source_row}: {row.name} | job `{row.job_name}` | dept `{row.department_name}` | {row.review_note}"
        )

    return "\n".join(sections) + "\n"


def validate_outputs(cleaned_rows: list[CleanedRow]) -> None:
    assert len(cleaned_rows) == 94, f"Expected 94 employee rows, got {len(cleaned_rows)}"
    assert all(row.name == row.name.strip() for row in cleaned_rows)
    assert all(row.job_key in JOB_SPECS for row in cleaned_rows)
    assert all(row.department_key in DEPARTMENTS for row in cleaned_rows)
    assert all(not row.work_email or "@" in row.work_email for row in cleaned_rows)
    employee_ids = {row.employee_id for row in cleaned_rows}
    assert all(not row.manager_employee_id or row.manager_employee_id in employee_ids for row in cleaned_rows)
    assert all(
        (not row.work_location)
        or (row.location_partner_id)
        or ("multiple locations:" in row.review_note)
        for row in cleaned_rows
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    source_rows = collect_rows(args.workbook)
    cleaned_rows = clean_rows(source_rows)
    validate_outputs(cleaned_rows)
    outputs = export_outputs(cleaned_rows, args.output_dir)

    print(f"Processed {len(cleaned_rows)} employee rows.")
    for label, path in outputs.items():
        print(f"{label}: {path}")


if __name__ == "__main__":
    main()
