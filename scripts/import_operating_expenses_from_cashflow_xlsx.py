#!/usr/bin/env python3
from __future__ import annotations

import argparse
from calendar import monthrange
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from odoo import Command


PIVOT_SHEET_MONTHS = {
    "pvt Jan 26": 1,
    "pvt Feb 26": 2,
    "Pvt Mar 26": 3,
}
TECHNICAL_EMPLOYEE_NAME = "TENENET Prevádzkové náklady Import"


def normalize_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def parse_amount(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        text = normalize_text(value).replace(" ", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0


def month_end(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def _find_tnnt_columns(sheet):
    for row_idx in (1, 2, 3):
        for col_idx in range(1, min(sheet.max_column, 20) + 1):
            if normalize_text(sheet.cell(row=row_idx, column=col_idx).value).casefold() == "tnnt":
                return col_idx, col_idx + 1
    raise ValueError(f"TNNT block not found in sheet {sheet.title}")


def _iter_tnnt_operating_rows(sheet):
    label_col, amount_col = _find_tnnt_columns(sheet)
    start_row = None
    for row_idx in range(1, sheet.max_row + 1):
        label = normalize_text(sheet.cell(row=row_idx, column=label_col).value)
        if label.casefold() == "naklady prevadzkove":
            start_row = row_idx + 1
            break
    if not start_row:
        return

    for row_idx in range(start_row, sheet.max_row + 1):
        label = normalize_text(sheet.cell(row=row_idx, column=label_col).value)
        amount = parse_amount(sheet.cell(row=row_idx, column=amount_col).value)
        folded = label.casefold()
        if not label and not amount:
            continue
        if folded.startswith("naklady ") and folded != "naklady prevadzkove":
            break
        if folded.startswith("trzby") or folded == "grand total" or folded == "total" or folded == "check":
            break
        if amount >= 0.0:
            continue
        yield {
            "row_idx": row_idx,
            "label": label,
            "amount": abs(amount),
        }


def parse_operating_expense_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    for sheet_name, month in PIVOT_SHEET_MONTHS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for item in _iter_tnnt_operating_rows(sheet) or []:
            rows.append({
                **item,
                "sheet_name": sheet_name,
                "month": month,
            })
    return rows


def _get_or_create_technical_employee(env):
    employee = env["hr.employee"].with_context(active_test=False).search(
        [("name", "=", TECHNICAL_EMPLOYEE_NAME)],
        limit=1,
    )
    if employee:
        return employee
    return env["hr.employee"].create({
        "name": TECHNICAL_EMPLOYEE_NAME,
        "work_ratio": 100.0,
    })


def _build_source_key(workbook_path: Path, sheet_name: str, row_idx: int):
    return f"operating_cf:{workbook_path.name}:{sheet_name}:{row_idx}"


def import_operating_expenses(env, workbook_path: str | Path, year: int = 2026):
    workbook_path = Path(workbook_path)
    rows = parse_operating_expense_rows(workbook_path)
    employee = _get_or_create_technical_employee(env)
    config_model = env["tenenet.expense.type.config"]
    config_model._load_default_operating_seed_data()
    tax = config_model._get_default_operating_tax()
    HrExpense = env["hr.expense"].with_context(mail_create_nosubscribe=True)

    created = 0
    updated = 0
    by_category = Counter()

    for row in rows:
        config = config_model._find_or_create_operating_type_for_detail_label(row["label"])
        source_key = _build_source_key(workbook_path, row["sheet_name"], row["row_idx"])
        notes = (
            f"Import z workbooku {workbook_path.name}, hárok {row['sheet_name']}, riadok {row['row_idx']}. "
            f"Pôvodný label: {row['label']}"
        )
        values = {
            "name": row["label"],
            "description": notes,
            "employee_id": employee.id,
            "date": month_end(year, row["month"]),
            "total_amount_currency": row["amount"],
            "payment_mode": "company_account",
            "tax_ids": [Command.set(tax.ids)],
            "tenenet_cost_flow": "operating",
            "tenenet_expense_type_config_id": config.id,
            "tenenet_import_source_key": source_key,
        }
        existing = HrExpense.search([("tenenet_import_source_key", "=", source_key)], limit=1)
        if existing:
            existing.write(values)
            updated += 1
        else:
            HrExpense.create(values)
            created += 1
        by_category[config.display_name] += 1

    return {
        "created": created,
        "updated": updated,
        "rows": len(rows),
        "by_category": dict(by_category),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()

    if "env" not in globals():
        raise SystemExit("Run this script inside `odoo shell` so `env` is available.")

    result = import_operating_expenses(env, args.workbook, year=args.year)
    env.cr.commit()
    print(result)


if __name__ == "__main__":
    main()
