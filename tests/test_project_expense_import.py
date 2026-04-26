import importlib.util
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from odoo.tests import TransactionCase, tagged


def _load_module(module_name: str, file_path: Path):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader
    sys.modules[module_name] = module
    try:
        spec.loader.exec_module(module)
    except Exception:
        sys.modules.pop(module_name, None)
        raise
    return module


@tagged("post_install", "-at_install")
class TestProjectExpenseImport(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        scripts_dir = Path(__file__).resolve().parents[1] / "scripts"
        cls.import_script = _load_module(
            "tenenet_import_project_expenses_from_cashflow_plan_xlsx",
            scripts_dir / "import_project_expenses_from_cashflow_plan_xlsx.py",
        )

    def setUp(self):
        super().setUp()
        self.program = self.env["tenenet.program"].create({"name": "SCPP", "code": "SCPP_TEST"})
        self.project = self.env["tenenet.project"].create({
            "name": "ICM",
            "program_ids": [(6, 0, self.program.ids)],
        })

    def test_import_creates_project_linked_hr_expenses_and_is_idempotent(self):
        with TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "cashflow.xlsx"
            self._write_workbook(workbook_path)

            first = self.import_script.import_project_expenses(self.env, workbook_path, year=2026)
            second = self.import_script.import_project_expenses(self.env, workbook_path, year=2026)

            expenses = self.env["hr.expense"].search([
                ("tenenet_import_source_key", "like", f"project_cf_plan:{workbook_path.name}:%"),
            ])

            self.assertEqual(first["created"], 2)
            self.assertEqual(first["updated"], 0)
            self.assertEqual(first["skipped"], 1)
            self.assertEqual(second["created"], 0)
            self.assertEqual(second["updated"], 2)
            self.assertEqual(len(expenses), 2)
            self.assertTrue(all(expense.tenenet_cost_flow == "project" for expense in expenses))
            self.assertTrue(all(expense.tenenet_project_id == self.project for expense in expenses))
            self.assertTrue(all(expense.tenenet_project_expense_ids for expense in expenses))
            self.assertTrue(all(expense.tax_ids.filtered(lambda tax: tax.amount == 23.0) for expense in expenses))

            by_month = {expense.date.month: expense for expense in expenses}
            self.assertEqual(set(by_month), {4, 5})
            self.assertEqual(
                by_month[4].tenenet_expense_type_config_id.cashflow_row_key,
                "workbook:expense:projektove-naklady-icm",
            )
            self.assertEqual(
                by_month[4].tenenet_expense_type_config_id.cashflow_row_label,
                "Projektové náklady - ICM",
            )
            self.assertEqual(
                first["skipped_labels"],
                {"Projektovy naklad - Guide, Stem, MinM, EASY": 1},
            )

    def _write_workbook(self, path: Path):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CF 2026 (rolling)"
        sheet.cell(row=57, column=21, value="Projektove naklady - ICM")
        sheet.cell(row=57, column=25, value=-725.0)
        sheet.cell(row=57, column=26, value=-725.0)
        sheet.cell(row=59, column=21, value="Projektovy naklad - Guide, Stem, MinM, EASY")
        sheet.cell(row=59, column=25, value=-1500.0)
        workbook.save(path)
