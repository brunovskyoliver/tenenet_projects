import importlib.util
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from odoo.tests import TransactionCase, tagged


def _load_module(module_name: str, file_path: Path):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader
    sys.modules[module_name] = module
    try:
        spec.loader.exec_module(module)
    except Exception:
        sys.modules.pop(module_name, None)
        raise
    return module


@tagged("post_install", "-at_install")
class TestOperatingExpenseImport(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        scripts_dir = Path(__file__).resolve().parents[1] / "scripts"
        cls.import_script = _load_module(
            "tenenet_import_operating_expenses_from_cashflow_xlsx",
            scripts_dir / "import_operating_expenses_from_cashflow_xlsx.py",
        )

    def test_import_creates_hr_expenses_and_is_idempotent(self):
        with TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "cashflow.xlsx"
            self._write_workbook(workbook_path)

            first = self.import_script.import_operating_expenses(self.env, workbook_path, year=2026)
            second = self.import_script.import_operating_expenses(self.env, workbook_path, year=2026)

            expenses = self.env["hr.expense"].search([
                ("tenenet_import_source_key", "like", f"operating_cf:{workbook_path.name}:%"),
            ])
            labels = set(expenses.mapped("name"))
            source_keys = set(expenses.mapped("tenenet_import_source_key"))

            self.assertEqual(first["created"], 3)
            self.assertEqual(first["updated"], 0)
            self.assertEqual(second["created"], 0)
            self.assertEqual(second["updated"], 3)
            self.assertEqual(len(expenses), 3)
            self.assertEqual(len(source_keys), 3)
            self.assertEqual(labels, {"Najom", "IT slu", "PSC"})
            self.assertTrue(all(expense.tenenet_cost_flow == "operating" for expense in expenses))
            self.assertTrue(all(expense.tenenet_internal_expense_ids for expense in expenses))
            self.assertTrue(all(expense.tax_ids for expense in expenses))
            self.assertTrue(all(expense.tax_ids.filtered(lambda tax: tax.amount == 23.0) for expense in expenses))

            by_name = {expense.name: expense for expense in expenses}
            self.assertEqual(by_name["Najom"].tenenet_expense_type_config_id.seed_key, "operating_rent")
            self.assertEqual(by_name["IT slu"].tenenet_expense_type_config_id.seed_key, "operating_it")
            self.assertEqual(by_name["PSC"].tenenet_expense_type_config_id.seed_key, "operating_other")

    def _write_workbook(self, path: Path):
        from openpyxl import Workbook

        workbook = Workbook()
        jan = workbook.active
        jan.title = "pvt Jan 26"
        jan.cell(row=1, column=1, value="TNNT")
        jan.cell(row=3, column=1, value="Row Labels")
        jan.cell(row=3, column=2, value="Sum of Suma")
        jan.cell(row=21, column=1, value="Naklady prevadzkove")
        jan.cell(row=22, column=1, value="Najom")
        jan.cell(row=22, column=2, value=-2480.96)
        jan.cell(row=23, column=1, value="IT slu")
        jan.cell(row=23, column=2, value=-616.02)
        jan.cell(row=24, column=1, value="Naklady projektove")
        jan.cell(row=25, column=1, value="Projekt X")
        jan.cell(row=25, column=2, value=-10.0)

        mar = workbook.create_sheet("Pvt Mar 26")
        mar.cell(row=1, column=1, value="TNNT")
        mar.cell(row=2, column=1, value="Row Labels")
        mar.cell(row=2, column=2, value="Sum of Suma")
        mar.cell(row=23, column=1, value="Naklady prevadzkove")
        mar.cell(row=24, column=1, value="PSC")
        mar.cell(row=24, column=2, value=-5782.4)
        mar.cell(row=25, column=1, value="Trzby")
        mar.cell(row=25, column=2, value=100.0)

        workbook.save(path)
