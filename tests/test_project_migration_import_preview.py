import importlib.util
import sys
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from odoo.tests import TransactionCase, tagged


def _load_module(module_name: str, file_path: Path):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader
    sys.modules[module_name] = module
    try:
        spec.loader.exec_module(module)
    except Exception:
        sys.modules.pop(module_name, None)
        raise
    return module


@tagged("post_install", "-at_install")
class TestProjectMigrationImportPreview(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        scripts_dir = Path(__file__).resolve().parents[1] / "scripts"
        cls.import_script = _load_module(
            "tenenet_import_project_migration_workbooks",
            scripts_dir / "import_project_migration_workbooks.py",
        )

    def setUp(self):
        super().setUp()
        self.program = self.env["tenenet.program"].with_context(active_test=False).search(
            [("code", "=", "SPODASK")],
            limit=1,
        )
        self.existing_project = self.env["tenenet.project"].create({
            "name": "EASPD GUIDE",
            "project_type": "narodny",
            "program_ids": [(6, 0, self.program.ids)],
        })
        self.employee = self.env["hr.employee"].create({
            "name": "Dugasová Kvetoslava",
            "work_ratio": 100.0,
        })

    def test_build_preview_parses_expected_sheets_and_writes_report(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            report = tmp_path / "report.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments)
            self._write_cashflow(cashflow)

            result = self.import_script.build_migration_preview(
                self.env,
                project_summary,
                assignments,
                cashflow,
                report,
            )

            self.assertTrue(report.exists())
            self.assertEqual(len(result["projects"]), 14)
            self.assertEqual(len(result["assignments"]), 1)
            self.assertEqual(len(result["timesheets"]), 3)
            self.assertEqual(len(result["cashflows"]), 1)
            self.assertEqual(len(result["cashflow_plan_rows"]), 1)

            workbook = self._load_report(report)
            self.assertEqual(
                set(workbook.sheetnames),
                {"Projects", "Assignments", "Timesheets", "Cashflows", "Cashflow Plan", "Unmatched", "Warnings", "Summary", "Employee Matches", "Employee Monthly Summaries"},
            )

    def test_project_program_aliases_and_unknown_warning(self):
        with TemporaryDirectory() as tmp_dir:
            project_summary = Path(tmp_dir) / "project_summary.xlsx"
            self._write_project_summary(project_summary)

            warnings = []
            projects = self.import_script.parse_project_summary(self.env, project_summary, warnings)
            by_name = {project.project_name: project for project in projects}

            self.assertIn("ID-CITIZEN", by_name)
            self.assertEqual(by_name["ID-CITIZEN"].project_type, "medzinarodny")
            self.assertEqual(by_name["SPOD alias"].program_code, "SPODASK")
            self.assertEqual(by_name["NAS alias"].program_code, "NAS_A_VAZ")
            self.assertEqual(by_name["VCI alias"].program_code, "VCI")
            self.assertEqual(by_name["SSP alias"].program_code, "SSP")
            self.assertTrue(any(w.message == "Unknown project program" and w.raw_value == "Mystery" for w in warnings))

    def test_assignment_parser_skips_aggregate_and_total_columns(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments)

            warnings = []
            unmatched = []
            projects = self.import_script.parse_project_summary(self.env, project_summary, warnings)
            assignment_rows, timesheet_rows = self.import_script.parse_assignments(
                self.env,
                assignments,
                self.import_script.build_project_index(projects),
                unmatched,
                warnings,
            )

            self.assertEqual(len(assignment_rows), 1)
            self.assertEqual(len(timesheet_rows), 3)
            self.assertEqual(sum(row.hours_pp for row in timesheet_rows), 12.0)
            self.assertNotIn("summary_Tenenet", {row.sheet for row in timesheet_rows})
            self.assertEqual(assignment_rows[0].employee_id, self.employee.id)
            self.assertEqual(assignment_rows[0].employee_match_status, "odoo")
            self.assertAlmostEqual(assignment_rows[0].wage_ccp, 50.0, places=2)
            self.assertAlmostEqual(assignment_rows[0].contribution_multiplier, 1.362, places=4)
            self.assertEqual(assignment_rows[0].date_end, date(2026, 12, 31))
            self.assertEqual([row.ccp_amount for row in timesheet_rows], [200.0, 200.0, 200.0])

    def test_import_updates_employee_multiplier_from_employee_sheet(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            report = tmp_path / "report.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments, employee_sheet_multiplier=1.3045)
            self._write_cashflow(cashflow)

            result = self.import_script.build_migration_preview(
                self.env,
                project_summary,
                assignments,
                cashflow,
                report,
            )
            self.import_script.apply_migration_preview(self.env, result)

            employee = self.env["hr.employee"].browse(self.employee.id)
            assignment = self.env["tenenet.project.assignment"].search([
                ("employee_id", "=", self.employee.id),
                ("project_id", "=", self.existing_project.id),
            ], limit=1)

            self.assertAlmostEqual(employee.tenenet_payroll_contribution_multiplier, 1.3045, places=4)
            self.assertAlmostEqual(assignment.wage_ccp, 50.0, places=2)
            self.assertAlmostEqual(assignment.wage_hm, 50.0 / 1.3045, places=4)

    def test_import_updates_employee_monthly_summary_from_employee_sheet(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            report = tmp_path / "report.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments, employee_sheet_summary=True)
            self._write_cashflow(cashflow)

            result = self.import_script.build_migration_preview(
                self.env,
                project_summary,
                assignments,
                cashflow,
                report,
            )
            self.import_script.apply_migration_preview(self.env, result)

            march_cost = self.env["tenenet.employee.tenenet.cost"].search([
                ("employee_id", "=", self.employee.id),
                ("period", "=", date(2026, 3, 1)),
            ], limit=1)

            self.assertTrue(march_cost.imported_from_migration_workbook)
            self.assertAlmostEqual(march_cost.imported_capacity_hours_incl, 176.0, places=2)
            self.assertAlmostEqual(march_cost.imported_capacity_hours, 168.0, places=2)
            self.assertAlmostEqual(march_cost.imported_total_gross_salary, 6400.0, places=2)
            self.assertAlmostEqual(march_cost.imported_total_labor_cost, 8348.8, places=2)
            self.assertAlmostEqual(march_cost.imported_worked_hours, 120.0, places=2)
            self.assertAlmostEqual(march_cost.imported_holidays_hours, 8.0, places=2)
            self.assertAlmostEqual(march_cost.imported_vacation_hours, 16.0, places=2)
            self.assertAlmostEqual(march_cost.imported_doctor_hours, 4.0, places=2)
            self.assertAlmostEqual(march_cost.imported_internal_gross_salary, 1500.0, places=2)
            self.assertAlmostEqual(march_cost.imported_internal_labor_cost, 1950.0, places=2)

    def test_timesheet_labor_cost_override_matches_imported_ccp(self):
        assignment = self.env["tenenet.project.assignment"].with_context(
            skip_tenenet_assignment_capacity_check=True,
        ).create({
            "employee_id": self.employee.id,
            "project_id": self.existing_project.id,
            "date_start": date(2026, 1, 1),
            "date_end": date(2026, 12, 31),
            "allocation_ratio": 100.0,
            "wage_hm": 16.8063,
        })
        timesheet = self.env["tenenet.project.timesheet"]._get_or_create_for_assignment_period(
            assignment,
            date(2026, 3, 1),
        )

        timesheet.write({
            "hours_pp": 123.2,
            "labor_cost_override": 2734.61,
        })

        self.assertAlmostEqual(timesheet.total_labor_cost, 2734.61, places=2)

    def test_assignment_parser_uses_source_workbook_employee_list_as_fallback(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments, employee_name="Workbook Person", include_employee_list=True)

            warnings = []
            unmatched = []
            projects = self.import_script.parse_project_summary(self.env, project_summary, warnings)
            assignment_rows, _timesheet_rows = self.import_script.parse_assignments(
                self.env,
                assignments,
                self.import_script.build_project_index(projects),
                unmatched,
                warnings,
            )

            self.assertEqual(len(assignment_rows), 1)
            self.assertEqual(assignment_rows[0].employee_match_status, "source_workbook")
            self.assertFalse(any(row.kind == "employee" and row.key == "Workbook Person" for row in unmatched))

    def test_employee_aliases_ignore_titles_prefixes_and_suffixes(self):
        aliases = self.import_script.employee_aliases("Mgr. PhD. Nina Jašeková")
        self.assertIn("nina jasekova", aliases)
        self.assertIn("jasekova nina", aliases)

        aliases = self.import_script.employee_aliases("DVP_Dužďová Monika")
        self.assertIn("duzdova monika", aliases)
        self.assertIn("monika duzdova", aliases)

        aliases = self.import_script.employee_aliases("Tóthová Melitta 2")
        self.assertIn("tothova melitta", aliases)
        self.assertIn("melitta tothova", aliases)

    def test_employee_resolution_prefers_odoo_match_over_exact_source_alias(self):
        source_match = self.import_script.EmployeeMatch(
            name="Jašeková Nina 2h",
            match_status="source_workbook",
        )
        odoo_match = self.import_script.EmployeeMatch(
            name="Nina Jašeková",
            employee_id=191,
            match_status="odoo",
            source_ref="hr.employee:191",
        )
        index = {}
        self.import_script.add_employee_index_entry(index, odoo_match)
        self.import_script.add_employee_index_entry(index, source_match)

        self.assertEqual(
            self.import_script.resolve_employee(index, "Jašeková Nina 2h").employee_id,
            191,
        )

    def test_cashflow_parser_checks_month_total_and_reports_unmatched(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            self._write_project_summary(project_summary)
            self._write_cashflow(cashflow)

            warnings = []
            unmatched = []
            projects = self.import_script.parse_project_summary(self.env, project_summary, warnings)
            cashflow_rows = self.import_script.parse_cashflows(
                self.env,
                cashflow,
                "CF 2026",
                self.import_script.build_project_index(projects),
                unmatched,
                warnings,
            )

            self.assertEqual(len(cashflow_rows), 1)
            self.assertAlmostEqual(sum(cashflow_rows[0].month_amounts.values()), cashflow_rows[0].receipt_amount, places=2)
            self.assertTrue(any(row.kind == "project" and row.key == "999 Missing project" for row in unmatched))
            self.assertFalse(any(row.kind == "cashflow" and row.key == "Cash-OUT Bank Statement" for row in unmatched))

    def test_cashflow_plan_parser_reads_right_table_fallback_labels(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            self._write_project_summary(project_summary)
            self._write_cashflow(cashflow)

            warnings = []
            projects = self.import_script.parse_project_summary(self.env, project_summary, warnings)
            plan_rows = self.import_script.parse_cashflow_plan_rows(
                self.env,
                cashflow,
                "CF 2026",
                self.import_script.build_project_index(projects),
                warnings,
            )

            self.assertEqual(len(plan_rows), 1)
            self.assertEqual(plan_rows[0].row_label, "Prevadzkove N - PSC")
            self.assertEqual(plan_rows[0].row_type, "expense")
            self.assertEqual(plan_rows[0].row_key, "workbook:expense:prevadzkove-n-psc")
            self.assertAlmostEqual(plan_rows[0].month_amounts[3], -5782.4, places=2)

    def test_unmatched_employee_is_reported(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments, employee_name="Neznáma Osoba", include_employee_list=False)

            warnings = []
            unmatched = []
            projects = self.import_script.parse_project_summary(self.env, project_summary, warnings)
            self.import_script.parse_assignments(
                self.env,
                assignments,
                self.import_script.build_project_index(projects),
                unmatched,
                warnings,
            )

            self.assertTrue(any(row.kind == "employee" and row.key == "Neznáma Osoba" for row in unmatched))

    def test_missing_assignment_project_becomes_project_candidate(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            report = tmp_path / "report.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments, extra_project_sheet="summary_New Missing")
            self._write_cashflow(cashflow)

            result = self.import_script.build_migration_preview(
                self.env,
                project_summary,
                assignments,
                cashflow,
                report,
            )

            by_name = {project.project_name: project for project in result["projects"]}
            self.assertIn("New Missing", by_name)
            self.assertEqual(by_name["New Missing"].match_status, "create_from_assignment_workbook")
            self.assertEqual(by_name["New Missing"].contract_number, "New Missing")
            self.assertEqual(by_name["New Missing"].description, "")
            self.assertEqual(by_name["New Missing"].project_type, "medzinarodny")
            self.assertFalse(any(
                row.source == "Assignments" and row.kind == "project" and row.key == "New Missing"
                for row in result["unmatched"]
            ))

    def test_missing_assignment_project_uses_project_program_binding_workbook(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            bindings = tmp_path / "bindings.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            report = tmp_path / "report.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(assignments, extra_project_sheet="summary_Bound Project")
            self._write_assignments(bindings, project_bindings={"Bound Project": "Admin"})
            self._write_cashflow(cashflow)

            result = self.import_script.build_migration_preview(
                self.env,
                project_summary,
                assignments,
                cashflow,
                report,
                project_program_bindings_path=bindings,
            )

            by_name = {project.project_name: project for project in result["projects"]}
            self.assertEqual(by_name["Bound Project"].program_raw, "Admin")
            self.assertEqual(by_name["Bound Project"].program_code, "ADMIN_TENENET")
            self.assertEqual(by_name["Bound Project"].project_type, "narodny")

    def test_apz_special_assignment_workbook_adds_ratio_preview_rows(self):
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            project_summary = tmp_path / "project_summary.xlsx"
            assignments = tmp_path / "assignments.xlsx"
            cashflow = tmp_path / "cashflow.xlsx"
            apz_assignments = tmp_path / "apz.xlsx"
            report = tmp_path / "report.xlsx"
            self._write_project_summary(project_summary)
            self._write_assignments(
                assignments,
                include_employee_list=True,
                employee_names=["Sandberg Svetlana", "Filkász Gabriel"],
                extra_project_sheet="summary_APZ_2N",
            )
            self._write_cashflow(cashflow)
            self._write_apz_special(apz_assignments)

            result = self.import_script.build_migration_preview(
                self.env,
                project_summary,
                assignments,
                cashflow,
                report,
                apz_assignments_path=apz_assignments,
            )

            apz_rows = [row for row in result["assignments"] if row.sheet == "apz_special:Sheet 1"]
            self.assertEqual(len(apz_rows), 2)
            by_employee = {row.employee_name: row for row in apz_rows}
            self.assertEqual(by_employee["Sandberg Svetlana"].monthly_ratios[1], 50.0)
            self.assertEqual(by_employee["Filkász Gabriel"].monthly_ratios[4], 50.0)
            self.assertFalse(any(row.source == "APZ Special Assignments" and row.kind == "project" for row in result["unmatched"]))

    def _write_project_summary(self, path: Path) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Projects Summary MG"
        headers = {
            1: "Rok",
            3: "Project",
            4: "Name",
            6: "Number",
            7: "Recipient",
            8: "Project Type",
            9: "Donor",
            10: "Program",
            20: "Project Budget",
            26: "Start",
            27: "End",
            34: "To be received in 2026",
        }
        for column, value in headers.items():
            sheet.cell(row=2, column=column, value=value)

        names = [
            ("EASPD GUIDE", "Guide project", "SPODaSK"),
            ("SPOD alias", "SPOD alias description", "SPODaSK"),
            ("NAS alias", "NAS alias description", "Násilie a väzenstvo"),
            ("VCI alias", "VCI alias description", "VCI"),
            ("SSP alias", "SSP alias description", "ŠSP"),
            ("Unknown program", "Unknown program description", "Mystery"),
            ("Project 07", "Description 07", "SPODaSK"),
            ("Project 08", "Description 08", "SPODaSK"),
            ("Project 09", "Description 09", "SPODaSK"),
            ("Project 10", "Description 10", "SPODaSK"),
            ("Project 11", "Description 11", "SPODaSK"),
            ("Project 12", "Description 12", "SPODaSK"),
            ("Project 13", "Description 13", "SPODaSK"),
            ("ID-CITIZEN", "ID: CITIZEN - Solving Accessibility Challenges", "Vyšehradský fond"),
        ]
        for offset, (project, description, program) in enumerate(names, start=3):
            sheet.cell(row=offset, column=1, value="2025-2026" if project == "ID-CITIZEN" else 2026)
            sheet.cell(row=offset, column=3, value=project)
            sheet.cell(row=offset, column=4, value=description)
            sheet.cell(row=offset, column=6, value=f"CN-{offset}")
            sheet.cell(row=offset, column=7, value="TENENET")
            sheet.cell(row=offset, column=8, value="11 Vysegrad Fund" if project == "ID-CITIZEN" else "1 SR - samospráva")
            sheet.cell(row=offset, column=9, value="Donor")
            sheet.cell(row=offset, column=10, value=program)
            sheet.cell(row=offset, column=20, value=1200)
            sheet.cell(row=offset, column=26, value=date(2025, 5, 20) if project == "ID-CITIZEN" else date(2026, 1, 1))
            sheet.cell(row=offset, column=27, value=date(2026, 5, 20) if project == "ID-CITIZEN" else date(2026, 12, 31))
            sheet.cell(row=offset, column=34, value=1200)
        workbook.save(path)

    def _write_assignments(
        self,
        path: Path,
        employee_name: str = "Dugasová Kvetoslava",
        include_employee_list: bool = False,
        employee_names: list[str] | None = None,
        extra_project_sheet: str | None = None,
        project_bindings: dict[str, str] | None = None,
        employee_sheet_multiplier: float | None = None,
        employee_sheet_summary: bool = False,
    ) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        template = workbook.active
        template.title = "Meno Priezvisko - template"
        for column in [3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15]:
            template.cell(row=4, column=column, value=160)

        if project_bindings is not None:
            project_list = workbook.create_sheet("Zoznam projektov")
            project_list.cell(row=1, column=2, value="Projekt")
            for sequence, (project_name, program_name) in enumerate(project_bindings.items(), start=2):
                project_list.cell(row=sequence, column=1, value=sequence - 1)
                project_list.cell(row=sequence, column=2, value=project_name)
                project_list.cell(row=sequence, column=3, value=program_name)

        if include_employee_list:
            employee_names = employee_names or [employee_name]
            employee_list = workbook.create_sheet("Zoznam zamestnancov")
            employee_list.append(["Číslo", "Titul", "Zamestnanec", "Priezvisko", "Meno"])
            for sequence, source_employee_name in enumerate(employee_names, start=1):
                employee_list.append([
                    sequence,
                    "",
                    source_employee_name,
                    source_employee_name.split()[0],
                    " ".join(source_employee_name.split()[1:]),
                ])

        aggregate = workbook.create_sheet("summary_Tenenet")
        aggregate.cell(row=1, column=1, value="Name")
        aggregate.cell(row=2, column=1, value="Aggregate Person")
        aggregate.cell(row=2, column=2, value="Celkova cena prace")
        aggregate.cell(row=2, column=3, value=1000)
        aggregate.cell(row=3, column=1, value="Aggregate Person")
        aggregate.cell(row=3, column=2, value="Odpracovane hodiny")
        aggregate.cell(row=3, column=3, value=100)

        sheet = workbook.create_sheet("summary_EASPD GUIDE")
        sheet.cell(row=1, column=1, value="Name")
        sheet.cell(row=1, column=2, value="Column2")
        for column, label in {3: "1 26", 4: "2 26", 5: "3 26", 9: "1/2 Total", 16: "2/2 Total 2026"}.items():
            sheet.cell(row=1, column=column, value=label)
        sheet.cell(row=2, column=1, value=employee_name)
        sheet.cell(row=2, column=2, value="Celkova cena prace")
        sheet.cell(row=2, column=3, value=200)
        sheet.cell(row=2, column=4, value=200)
        sheet.cell(row=2, column=5, value=200)
        sheet.cell(row=2, column=9, value=9999)
        sheet.cell(row=2, column=16, value=9999)
        sheet.cell(row=3, column=1, value=employee_name)
        sheet.cell(row=3, column=2, value="Odpracovane hodiny")
        sheet.cell(row=3, column=3, value=4)
        sheet.cell(row=3, column=4, value=4)
        sheet.cell(row=3, column=5, value=4)
        sheet.cell(row=3, column=9, value=9999)
        sheet.cell(row=3, column=16, value=9999)

        if employee_sheet_multiplier is not None or employee_sheet_summary:
            employee_sheet = workbook.create_sheet(employee_name)
            if employee_sheet_multiplier is not None:
                employee_sheet.cell(row=17, column=2, value="updatnut podla skutocnosti")
                for column in [3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15]:
                    employee_sheet.cell(row=17, column=column, value=employee_sheet_multiplier)
            if employee_sheet_summary:
                month_columns = [3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15]
                for column, value in zip(month_columns, [176, 160, 176, 160, 168, 176, 184, 168, 176, 176, 168, 168]):
                    employee_sheet.cell(row=3, column=1, value="Hours/month (incl holidays)")
                    employee_sheet.cell(row=3, column=column, value=value)
                for column, value in zip(month_columns, [160, 160, 168, 160, 168, 176, 176, 168, 176, 176, 168, 160]):
                    employee_sheet.cell(row=4, column=1, value="Hours/month")
                    employee_sheet.cell(row=4, column=column, value=value)
                for row, label, values in [
                    (6, "Hruba Mzda", [6400, 6400, 6400, 6400, 6400, 6400, 6400, 6400, 6400, 6400, 6400, 6400]),
                    (8, "Celkova cena prace", [8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8, 8348.8]),
                    (9, "Odpracovane hodiny", [100, 110, 120, 130, 140, 150, 160, 170, 180, 190, 200, 210]),
                    (10, "Platené sviatky - hodiny", [0, 0, 8, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
                    (11, "Dovolenka Hodiny", [0, 0, 16, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
                    (12, "LEKAR - hodiny", [0, 0, 4, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
                ]:
                    employee_sheet.cell(row=row, column=2, value=label)
                    for column, value in zip(month_columns, values):
                        employee_sheet.cell(row=row, column=column, value=value)
                for row, first_label, second_label, values in [
                    (44, "Tenenet", "Hruba Mzda", [1200, 1300, 1500, 1600, 1700, 1800, 1900, 2000, 2100, 2200, 2300, 2400]),
                    (46, "Tenenet", "Celkova cena prace", [1560, 1690, 1950, 2080, 2210, 2340, 2470, 2600, 2730, 2860, 2990, 3120]),
                ]:
                    employee_sheet.cell(row=row, column=1, value=first_label)
                    employee_sheet.cell(row=row, column=2, value=second_label)
                    for column, value in zip(month_columns, values):
                        employee_sheet.cell(row=row, column=column, value=value)

        if extra_project_sheet:
            extra = workbook.create_sheet(extra_project_sheet)
            extra.cell(row=1, column=1, value="Name")
            extra.cell(row=1, column=2, value="Column2")
            extra.cell(row=2, column=1, value=employee_name)
            extra.cell(row=2, column=2, value="Celkova cena prace")
            extra.cell(row=2, column=3, value=100)
            extra.cell(row=3, column=1, value=employee_name)
            extra.cell(row=3, column=2, value="Odpracovane hodiny")
            extra.cell(row=3, column=3, value=5)
        workbook.save(path)

    def _write_cashflow(self, path: Path) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CF 2026"
        sheet.cell(row=2, column=21, value="Project")
        for column, month in enumerate(range(1, 13), start=22):
            sheet.cell(row=2, column=column, value=month)
        sheet.cell(row=2, column=34, value="Total 2025")

        sheet.cell(row=3, column=1, value="Program")
        sheet.cell(row=3, column=2, value="EASPD GUIDE")
        sheet.cell(row=3, column=21, value="EASPD GUIDE")
        sheet.cell(row=3, column=22, value=100)
        sheet.cell(row=3, column=23, value=200)
        sheet.cell(row=3, column=24, value=300)
        sheet.cell(row=3, column=34, value=600)

        sheet.cell(row=4, column=21, value="999 Missing project")
        sheet.cell(row=4, column=22, value=50)
        sheet.cell(row=4, column=34, value=50)

        sheet.cell(row=5, column=21, value="Cash-OUT Bank Statement")
        sheet.cell(row=5, column=22, value=-50)
        sheet.cell(row=5, column=34, value=-50)

        sheet.cell(row=65, column=2, value="Prevadzkove N - PSC")
        sheet.cell(row=65, column=24, value=-5782.4)
        sheet.cell(row=65, column=34, value=-5782.4)
        workbook.save(path)

    def _write_apz_special(self, path: Path) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet 1"
        sheet.cell(row=2, column=1, value="aktivita")
        sheet.cell(row=2, column=3, value="pozícia")
        sheet.cell(row=2, column=4, value="VRR/MRR")
        sheet.cell(row=2, column=9, value="január 26")
        sheet.cell(row=2, column=12, value="apríl 26")
        sheet.cell(row=3, column=1, value="4P1")
        sheet.cell(row=3, column=3, value="Manažér")
        sheet.cell(row=3, column=4, value="VRR")
        sheet.cell(row=3, column=9, value="Sandberg 50%")
        sheet.cell(row=3, column=12, value="Filkász 50%")
        workbook.save(path)

    def _load_report(self, path: Path):
        from openpyxl import load_workbook

        return load_workbook(path, read_only=True, data_only=True)
